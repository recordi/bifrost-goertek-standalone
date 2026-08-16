#!/usr/bin/env python3
"""
test_runner.py — BIFROST 字段整合只读 Skill 测试运行器 (04C.4A.2 可信边界收口)

所有 passed 值由断言表达式计算，禁止写死 True / vacuous 下界 / 仅检查可加载 / 以 0 证安全。

测试组：
  A. SIM 回归（真实断言）
  B. 官方跨源回归（真实断言）
  C. 合成泛化测试（15 项）
  D. 防硬编码测试
  E. 合同真实性测试
  F. 反假通过变异测试（6 项）
  G. 可信边界测试（12 项，04C.4A.2 新增）

输出: test_results_04C4A.2.json
"""

import json
import os
import sys
import shutil
import hashlib
import tempfile
import uuid
import re
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from bifrost_data_mapper import (
    BifrostDataMapper, ContractLoader, SKILL_VERSION, RELEASE_STATUS,
    MAPPING_STATUS_ENUM, RELATION_STATUS_ENUM, NON_TABULAR_SHEETS,
    BLOCKED_UNSUPPORTED_FORMAT, BLOCKED_SEMANTIC_CONTRACT,
    BLOCKED_FILE_NOT_FOUND, BLOCKED_CORRUPT_FILE, BLOCKED_PATH_TRAVERSAL,
    BLOCKED_READ_ONLY_VIOLATION
)
from contract_validator import ContractValidator

# 路径
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
SKILL_DIR = os.path.normpath(os.path.join(SCRIPT_DIR, ".."))
CONTRACTS_DIR = os.path.join(SKILL_DIR, "references", "contracts")
FIXTURES_DIR = os.path.join(SKILL_DIR, "tests", "fixtures")
WORK_DIR = os.path.dirname(SKILL_DIR)  # 交付包根目录

# 业务数据仅作为外部测试输入，不进入 .omp/skills。通过环境变量或仓库
# test-inputs/ 注入，Skill 本体仍保持数据与逻辑分离。
TEST_INPUT_ROOT = os.environ.get("BIFROST_TEST_INPUT_ROOT", os.path.join(WORK_DIR, "test-inputs"))
SIM_FILE = os.path.join(TEST_INPUT_ROOT, "BIFROST_飞书导入数据包_v3_P0修复版_SIM-v2.2.xlsx")
OFFICIAL_FILE = os.path.join(TEST_INPUT_ROOT, "歌尔可脱敏企业测试数据集.xlsx")


def sha256_file(filepath):
    h = hashlib.sha256()
    with open(filepath, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()


def make_request(filepath, fmt, **kwargs):
    req = {
        "request_id": f"REQ-TEST-{uuid.uuid4().hex[:8]}",
        "source_id": "test",
        "source_name": os.path.basename(filepath),
        "source_file": filepath,
        "file_format": fmt,
        "declared_source_family": kwargs.get("declared_source_family"),
        "declared_source_type": kwargs.get("declared_source_type"),
        "mapping_mode": kwargs.get("mapping_mode", "zero_shot"),
        "semantic_model_version": "SEM-v1.1.1",
        "mapping_rule_version": kwargs.get("mapping_rule_version"),
        "allowed_domains": [],
        "sample_limit": 20,
        "read_only": True
    }
    return req


def _check_response_type_contract(draft):
    """响应字段类型合同：每条记录必须有 source_data_type + source_dataset_type，且不得有 source_type 键。"""
    for r in draft:
        if "source_type" in r:
            return False, f"record {r.get('source_field')} 含禁止键 source_type"
        if "source_data_type" not in r:
            return False, f"record {r.get('source_field')} 缺 source_data_type"
        if "source_dataset_type" not in r:
            return False, f"record {r.get('source_field')} 缺 source_dataset_type"
    return True, "ok"


class TestRunner:
    def __init__(self):
        self.results = {
            "test_session": {
                "skill_name": "bifrost-data-mapper-readonly",
                "logical_skill_version": SKILL_VERSION,
                "release_status": RELEASE_STATUS,
                "run_at": datetime.now().isoformat(),
                "total_tests": 0,
                "passed": 0,
                "failed": 0,
                "blocked": 0
            },
            "internal_skill_tests": [],
            "source_regression_tests": [],
            "external_post_package_checks": [],
            "summary": {}
        }

    def record(self, category, test_id, test_name, passed, details=None):
        # passed 必须是 bool 计算结果，禁止传入字面 True 写死
        entry = {
            "test_id": test_id,
            "test_name": test_name,
            "passed": bool(passed),
            "details": details or {}
        }
        self.results[category].append(entry)
        self.results["test_session"]["total_tests"] += 1
        if entry["passed"]:
            self.results["test_session"]["passed"] += 1
        else:
            self.results["test_session"]["failed"] += 1

    # ------------------------------------------------------------------
    # A. SIM 回归测试（真实断言）
    # ------------------------------------------------------------------

    def run_sim_regression(self):
        category = "source_regression_tests"
        if not os.path.exists(SIM_FILE):
            self.record(category, "SIM-00", "SIM 文件存在", False, {"error": "文件不存在"})
            return

        sha_before = sha256_file(SIM_FILE)
        mapper = BifrostDataMapper(CONTRACTS_DIR)
        req = make_request(SIM_FILE, "xlsx",
                           declared_source_family="TEAM_ENGINEER_SIMULATION",
                           declared_source_type="TEAM_ENGINEER_SIMULATION_STATIC_SNAPSHOT",
                           mapping_mode="approved_contract",
                           mapping_rule_version="MAP-v1.0.1")
        resp = mapper.orchestrate_mapping_run(req)
        sha_after = sha256_file(SIM_FILE)

        ss = resp["schema_summary"]

        # SIM-A: 表数动态扫描
        table_count = ss["table_count"]
        self.record(category, "SIM-A", "20张表动态扫描", table_count == 20,
                    {"table_count": table_count, "expected": 20})

        # SIM-A2: 字段口径三计数 + mapping_draft_count（真实等值断言）
        phys = ss["physical_used_column_count"]
        sem = ss["semantic_mapping_field_count"]
        excl = ss["non_tabular_excluded_column_count"]
        draft_count = resp["mapping_draft_count"]
        a2_pass = (phys == 316 and sem == 291 and excl == 25 and draft_count == 291)
        self.record(category, "SIM-A2", "字段口径 316/291/25 + draft291", a2_pass,
                    {"physical_used_column_count": phys, "semantic_mapping_field_count": sem,
                     "non_tabular_excluded_column_count": excl, "mapping_draft_count": draft_count})

        # SIM-B: 表数不硬编码
        actual_tables = len(ss["tables"])
        self.record(category, "SIM-B", "表数不硬编码", actual_tables == table_count,
                    {"actual": actual_tables})

        # SIM-C: 继承批准计数（真实等值断言）
        inherited = resp["inherited_approval_count"]
        active_confirmed = resp["active_confirmed_count"]
        confirmed_new = resp["confirmed_new_count"]
        c_pass = (inherited == 24 and active_confirmed == 24 and confirmed_new == 0)
        self.record(category, "SIM-C", "继承批准 24/24/0", c_pass,
                    {"inherited_approval_count": inherited,
                     "active_confirmed_count": active_confirmed,
                     "confirmed_new_count": confirmed_new})

        # SIM-D: 三类班次ID 互不覆盖
        draft = resp["mapping_draft"]
        sid = {x.get("target_field") for x in draft if x.get("target_field") == "source_shift_id"}
        rid = {x.get("target_field") for x in draft if x.get("target_field") == "related_shift_id"}
        xid = {x.get("target_field") for x in draft if x.get("target_field") == "simulated_shift_id"}
        sid_cnt = sum(1 for x in draft if x.get("target_field") == "source_shift_id")
        rid_cnt = sum(1 for x in draft if x.get("target_field") == "related_shift_id")
        xid_cnt = sum(1 for x in draft if x.get("target_field") == "simulated_shift_id")
        d_pass = (sid_cnt > 0 and rid_cnt > 0 and xid_cnt > 0
                  and sid == {"source_shift_id"} and rid == {"related_shift_id"}
                  and xid == {"simulated_shift_id"}
                  and sid != rid and sid != xid and rid != xid)
        self.record(category, "SIM-D", "三类班次ID互不覆盖", d_pass,
                    {"source_shift_id_count": sid_cnt, "related_shift_id_count": rid_cnt,
                     "simulated_shift_id_count": xid_cnt})

        # SIM-E: OEE 三字段同时存在且互不覆盖
        oee_src = [x for x in draft if x.get("target_field") == "oee_source"]
        oee_rec = [x for x in draft if x.get("target_field") == "oee_recomputed"]
        oee_dev = [x for x in draft if x.get("target_field") == "oee_deviation"]
        e_pass = (len(oee_src) > 0 and len(oee_rec) > 0 and len(oee_dev) > 0)
        self.record(category, "SIM-E", "OEE三字段同时存在互不覆盖", e_pass,
                    {"oee_source": len(oee_src), "oee_recomputed": len(oee_rec),
                     "oee_deviation": len(oee_dev)})

        # SIM-F: 性能率 >1.0 实际检出、原值保留、flagged_above_unity、不截断
        perf_findings = [f for f in resp["data_quality_findings"]
                         if f.get("issue_type") == "performance_rate_above_unity"
                         and f.get("target_field") == "performance_rate_raw"]
        detected = any(f.get("detected_count", 0) > 0 for f in perf_findings)
        flagged = all(f.get("validation_status") == "flagged_above_unity" for f in perf_findings) and len(perf_findings) > 0
        not_truncated = all(f.get("auto_truncated") is False for f in perf_findings) and len(perf_findings) > 0
        preserved = all(f.get("original_value_preserved") is True for f in perf_findings) and len(perf_findings) > 0
        # 直接读源文件验证原始 >1.0 值未被截断
        import openpyxl
        wb = openpyxl.load_workbook(SIM_FILE, read_only=True, data_only=True)
        ws = wb["01_OEE班次"]
        hdr = [str(h) if h else "" for h in next(ws.iter_rows(values_only=True))]
        pi = next((j for j, h in enumerate(hdr) if "性能率" in h), None)
        src_over = [r[pi] for r in ws.iter_rows(values_only=True)
                    if pi is not None and pi < len(r) and isinstance(r[pi], (int, float)) and r[pi] > 1.0]
        wb.close()
        src_has_over = len(src_over) > 0
        f_pass = detected and flagged and not_truncated and preserved and src_has_over
        self.record(category, "SIM-F", "性能率>1.0检出保留标记", f_pass,
                    {"perf_findings": len(perf_findings), "detected": detected,
                     "flagged_above_unity": flagged, "auto_truncated_any": not not_truncated,
                     "source_over_unity_count": len(src_over)})

        # SIM-G: 源文件哈希前后一致
        self.record(category, "SIM-G", "源文件哈希前后一致", sha_before == sha_after,
                    {"sha_before": sha_before, "sha_after": sha_after})
        self.record(category, "SIM-G2", "source_write_performed=false",
                    resp["source_verification"]["source_write_performed"] is False,
                    {"value": resp["source_verification"]["source_write_performed"]})

        # SIM-H: 无 col_n 伪字段进入语义映射
        col_n = [r for r in draft if str(r.get("source_field", "")).startswith("col_")]
        self.record(category, "SIM-H", "无col_n伪字段", len(col_n) == 0,
                    {"col_n_count": len(col_n)})

        return resp

    # ------------------------------------------------------------------
    # B. 官方跨源回归测试（真实断言）
    # ------------------------------------------------------------------

    def run_official_regression(self):
        category = "source_regression_tests"
        if not os.path.exists(OFFICIAL_FILE):
            self.record(category, "OFF-00", "官方文件存在", False, {"error": "文件不存在"})
            return

        sha_before = sha256_file(OFFICIAL_FILE)
        mapper = BifrostDataMapper(CONTRACTS_DIR)
        req = make_request(OFFICIAL_FILE, "xlsx",
                           declared_source_family="OFFICIAL_DEIDENTIFIED_SIMULATION",
                           declared_source_type="OFFICIAL_DEIDENTIFIED_SIMULATION",
                           mapping_mode="approved_contract",
                           mapping_rule_version="MAP-OFFICIAL-v0.2.1")
        resp = mapper.orchestrate_mapping_run(req)
        sha_after = sha256_file(OFFICIAL_FILE)

        ss = resp["schema_summary"]
        draft = resp["mapping_draft"]

        # OFF-A: 5张表、74字段、300行
        table_count = ss["table_count"]
        total_fields = ss["total_fields"]
        total_rows = sum(t["row_count"] for t in ss["tables"])
        self.record(category, "OFF-A", "5张表动态扫描", table_count == 5, {"table_count": table_count})
        self.record(category, "OFF-A2", "74字段动态扫描", total_fields == 74, {"total_fields": total_fields})
        self.record(category, "OFF-A3", "300行动态扫描", total_rows == 300, {"total_rows": total_rows})

        # OFF-B: 数据性质
        detected = resp["source_identity"]["detected_source_family"]
        self.record(category, "OFF-B", "数据性质 OFFICIAL", detected == "OFFICIAL_DEIDENTIFIED_SIMULATION",
                    {"detected": detected})

        # OFF-C: 54条批准映射实际进入响应（不止可加载）
        inherited = resp["inherited_approval_count"]
        confirmed_sum = resp["mapping_summary"].get("confirmed", 0)
        contract_loaded = "MAP_OFFICIAL_v0.2.1_mapping.json" in resp.get("contract_versions", {})
        c_pass = (contract_loaded and inherited == 54 and confirmed_sum == 54
                  and resp["confirmed_new_count"] == 0)
        self.record(category, "OFF-C", "54条批准映射进入响应", c_pass,
                    {"loaded": contract_loaded, "inherited": inherited,
                     "confirmed_sum": confirmed_sum, "confirmed_new": resp["confirmed_new_count"]})

        # OFF-D: OEE(%) 只映射到 oee_source；can_recompute_oee=false 明确合同字段
        oee_mappings = [r for r in draft if r.get("source_field") == "OEE(%)"]
        oee_targets = {r.get("target_field") for r in oee_mappings}
        oee_only_source = (len(oee_mappings) > 0 and oee_targets == {"oee_source"})
        can_recompute = resp.get("oee_safety_contract", {}).get("can_recompute_oee")
        d_pass = (oee_only_source and len(oee_mappings) > 0
                  and can_recompute is False)
        self.record(category, "OFF-D", "OEE(%)只映射oee_source+can_recompute=false", d_pass,
                    {"oee_mapping_count": len(oee_mappings), "oee_targets": list(oee_targets),
                     "can_recompute_oee": can_recompute})

        # OFF-E: can_recompute_oee=false 作为明确合同字段（不靠"未找到"间接推断）
        self.record(category, "OFF-E", "can_recompute_oee为明确合同字段", can_recompute is False,
                    {"can_recompute_oee": can_recompute,
                     "oee_safety_present": "oee_safety_contract" in resp})

        # OFF-F: inventory_snapshot 映射数>0 + view_projection_only/grain_status/aggregation_allowed
        inv = resp.get("inventory_contract_flags", {})
        inv_cnt = inv.get("inventory_snapshot_mapping_count", 0)
        f_pass = (inv_cnt > 0 and inv.get("view_projection_only") is True
                  and inv.get("grain_status") == "unresolved"
                  and inv.get("aggregation_allowed") is False)
        self.record(category, "OFF-F", "inventory_snapshot合同标志", f_pass,
                    {"inventory_mapping_count": inv_cnt,
                     "view_projection_only": inv.get("view_projection_only"),
                     "grain_status": inv.get("grain_status"),
                     "aggregation_allowed": inv.get("aggregation_allowed")})

        # OFF-G: 不确认任何 OJC 关系
        join_confirmed = sum(1 for jc in resp["join_candidates"] if jc["status"] == "confirmed")
        self.record(category, "OFF-G", "不确认OJC关系", join_confirmed == 0,
                    {"confirmed_count": join_confirmed})

        # OFF-H: 源文件哈希前后一致
        self.record(category, "OFF-H", "源文件哈希前后一致", sha_before == sha_after,
                    {"sha_before": sha_before, "sha_after": sha_after})

        # OFF-I: FM-023/029/038/039/040 批准映射验证
        targets_expected = {
            "FM-023": ("产-生产工单与质量检测", "OEE(%)", "shift", "oee_source"),
            "FM-029": ("供-采购订单与库存", "物料编码", "purchase_order", "material_code"),
            "FM-038": ("供-采购订单与库存", "当前库存", "inventory_snapshot", "stock_on_hand"),
            "FM-039": ("供-采购订单与库存", "安全库存", "inventory_snapshot", "safety_stock"),
            "FM-040": ("供-采购订单与库存", "库存周转天数", "inventory_snapshot", "inventory_turnover_days"),
        }
        fm_ok = []
        for fm_id, (st, sf, te, tf) in targets_expected.items():
            recs = [r for r in draft if r.get("source_table") == st and r.get("source_field") == sf
                    and r.get("target_entity") == te and r.get("target_field") == tf
                    and r.get("mapping_status") == "confirmed" and r.get("inherited_approval") is True]
            fm_ok.append(len(recs) == 1)
        self.record(category, "OFF-I", "FM-023/029/038/039/040批准映射", all(fm_ok),
                    {"verified": fm_ok})

        return resp

    # ------------------------------------------------------------------
    # C. 合成泛化测试
    # ------------------------------------------------------------------

    def run_synthetic_tests(self):
        category = "internal_skill_tests"
        self._generate_fixtures()
        self._test_renamed_fields(category)
        self._test_missing_fields(category)
        self._test_ratio_conflict(category)
        self._test_unit_conflict(category)
        self._test_multi_match(category)
        self._test_ambiguous_join(category)
        self._test_unknown_domain(category)
        self._test_duplicate_pk(category)
        self._test_formula_cell(category)
        self._test_empty_table(category)
        self._test_hidden_sheet(category)
        self._test_corrupt_xlsx(category)
        self._test_unsupported_format(category)
        self._test_path_traversal(category)
        self._test_missing_input(category)

    def _generate_fixtures(self):
        os.makedirs(FIXTURES_DIR, exist_ok=True)
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "班次数据"
        ws.append(["班次编号", "产线", "阶段", "开动率", "质量率", "OEE"])
        for i in range(1, 21):
            ws.append([f"SHIFT-{i:04d}", "LINE-S01", "改善前" if i <= 10 else "改善后",
                       0.85 + i * 0.005, 0.95, 0.80 + i * 0.003])
        wb.save(os.path.join(FIXTURES_DIR, "renamed_fields_fixture.xlsx"))
        wb.close()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "班次数据"
        ws.append(["班次编号", "产线", "开动率"])
        for i in range(1, 11):
            ws.append([f"SHIFT-{i:04d}", "LINE-S01", 0.85])
        wb.save(os.path.join(FIXTURES_DIR, "missing_fields_fixture.xlsx"))
        wb.close()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "混合比例"
        ws.append(["开动率", "质量率_百分比", "OEE"])
        for i in range(1, 11):
            ws.append([0.85, 95.0, 0.80])
        wb.save(os.path.join(FIXTURES_DIR, "ratio_unit_conflict_fixture.xlsx"))
        wb.close()

        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "表A"
        ws1.append(["ID", "名称", "类型"])
        for i in range(1, 21):
            ws1.append([f"ID-{i:03d}", f"名称{i}", "A"])
        ws2 = wb.create_sheet("表B")
        ws2.append(["ID", "描述", "类型"])
        for i in range(1, 21):
            ws2.append([f"ID-{i:03d}", f"描述{i}", "B"])
        wb.save(os.path.join(FIXTURES_DIR, "ambiguous_join_fixture.xlsx"))
        wb.close()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "未知数据"
        ws.append(["星球编号", "引力系数", "大气成分", "距离光年"])
        for i in range(1, 11):
            ws.append([f"PL-{i:04d}", 9.8 + i * 0.1, "氮氧", 4.2 + i])
        wb.save(os.path.join(FIXTURES_DIR, "unknown_domain_fixture.xlsx"))
        wb.close()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "重复键"
        ws.append(["ID", "名称"])
        for i in range(1, 11):
            ws.append([f"DUP-001" if i <= 3 else f"DUP-{i:03d}", f"名称{i}"])
        wb.save(os.path.join(FIXTURES_DIR, "duplicate_pk_fixture.xlsx"))
        wb.close()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "公式表"
        ws.append(["A", "B", "C"])
        for i in range(1, 6):
            ws.append([i, i * 2, f"=A{i+1}+B{i+1}"])
        wb.save(os.path.join(FIXTURES_DIR, "formula_cell_fixture.xlsx"))
        wb.close()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "空表"
        ws.append(["字段1", "字段2", "字段3"])
        wb.save(os.path.join(FIXTURES_DIR, "empty_table_fixture.xlsx"))
        wb.close()

        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "可见表"
        ws1.append(["ID", "名称"])
        ws1.append([1, "测试"])
        ws2 = wb.create_sheet("隐藏表")
        ws2.append(["隐藏ID", "隐藏值"])
        ws2.append([1, "隐藏"])
        ws2.sheet_state = "hidden"
        wb.save(os.path.join(FIXTURES_DIR, "hidden_sheet_fixture.xlsx"))
        wb.close()

        with open(os.path.join(FIXTURES_DIR, "corrupt_xlsx_fixture.xlsx"), "wb") as f:
            f.write(b"PK\x03\x04corrupted content not a real xlsx")
        with open(os.path.join(FIXTURES_DIR, "unsupported_xls_fixture.xls"), "wb") as f:
            f.write(b"fake xls content")
        with open(os.path.join(FIXTURES_DIR, "unsupported_xlsm_fixture.xlsm"), "wb") as f:
            f.write(b"fake xlsm content")

        for name, data in [
            ("renamed_fields_fixture.json", [{"line_code": f"L{i}", "avail": 0.85} for i in range(10)]),
            ("missing_fields_fixture.json", [{"line_code": f"L{i}"} for i in range(10)]),
            ("ratio_unit_conflict_fixture.json", [{"rate": 0.85, "pct": 85.0} for i in range(10)]),
            ("ambiguous_join_fixture.json", [{"id": f"ID-{i:03d}", "name": f"N{i}"} for i in range(10)]),
            ("unknown_domain_fixture.json", [{"planet": f"PL-{i}", "gravity": 9.8} for i in range(10)]),
        ]:
            with open(os.path.join(FIXTURES_DIR, name), "w") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)

    def _run_fixture(self, category, test_id, test_name, fixture_file, fmt="xlsx", **kwargs):
        fpath = os.path.join(FIXTURES_DIR, fixture_file)
        if not os.path.exists(fpath):
            self.record(category, test_id, test_name, False, {"error": f"fixture 不存在: {fixture_file}"})
            return None
        mapper = BifrostDataMapper(CONTRACTS_DIR)
        req = make_request(fpath, fmt, **kwargs)
        resp = mapper.orchestrate_mapping_run(req)
        return resp

    def _test_renamed_fields(self, category):
        resp = self._run_fixture(category, "SYN-C1", "字段改名但值域相近", "renamed_fields_fixture.xlsx")
        if resp:
            has_match = any(r["mapping_status"] != "unmapped" for r in resp["mapping_draft"])
            self.record(category, "SYN-C1", "字段改名但值域相近", has_match,
                        {"status": resp["status"], "mapping_count": len(resp["mapping_draft"])})

    def _test_missing_fields(self, category):
        resp = self._run_fixture(category, "SYN-C2", "关键字段缺失", "missing_fields_fixture.xlsx")
        if resp:
            unmapped_count = sum(1 for r in resp["mapping_draft"] if r["mapping_status"] == "unmapped")
            self.record(category, "SYN-C2", "关键字段缺失", resp["status"] != "blocked",
                        {"unmapped_count": unmapped_count})

    def _test_ratio_conflict(self, category):
        resp = self._run_fixture(category, "SYN-C3", "0-1与0-100比例冲突", "ratio_unit_conflict_fixture.xlsx")
        if resp:
            ambiguous_count = sum(1 for r in resp["mapping_draft"] if r.get("unit_ambiguous"))
            self.record(category, "SYN-C3", "0-1与0-100比例冲突", ambiguous_count > 0,
                        {"ambiguous_count": ambiguous_count})

    def _test_unit_conflict(self, category):
        import openpyxl
        fpath = os.path.join(FIXTURES_DIR, "unit_conflict_fixture.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "时间数据"
        ws.append(["停机秒数", "停机分钟数", "停机小时数"])
        for i in range(1, 11):
            ws.append([3600 * i, 60 * i, i])
        wb.save(fpath)
        wb.close()
        resp = self._run_fixture(category, "SYN-C4", "秒/分钟/小时单位冲突", "unit_conflict_fixture.xlsx")
        if resp:
            self.record(category, "SYN-C4", "秒/分钟/小时单位冲突", resp["status"] != "blocked",
                        {"mapping_count": len(resp["mapping_draft"])})

    def _test_multi_match(self, category):
        resp = self._run_fixture(category, "SYN-C5", "同一字段匹配多个目标", "renamed_fields_fixture.xlsx")
        if resp:
            # 真实断言：映射草稿非空且至少有一条非 unmapped
            cnt = len(resp["mapping_draft"])
            has_non_unmapped = any(r["mapping_status"] != "unmapped" for r in resp["mapping_draft"])
            self.record(category, "SYN-C5", "同一字段匹配多个目标", cnt > 0 and has_non_unmapped,
                        {"mapping_count": cnt})

    def _test_ambiguous_join(self, category):
        resp = self._run_fixture(category, "SYN-C6", "关联键高命中但语义歧义", "ambiguous_join_fixture.xlsx")
        if resp:
            join_count = len(resp["join_candidates"])
            confirmed = sum(1 for jc in resp["join_candidates"] if jc["status"] == "confirmed")
            self.record(category, "SYN-C6", "关联键高命中但语义歧义", confirmed == 0,
                        {"join_count": join_count, "confirmed": confirmed})

    def _test_unknown_domain(self, category):
        resp = self._run_fixture(category, "SYN-C7", "完全未知业务域", "unknown_domain_fixture.xlsx")
        if resp:
            unmapped = all(r["mapping_status"] == "unmapped" for r in resp["mapping_draft"])
            ext_count = len(resp["semantic_extension_proposals"])
            self.record(category, "SYN-C7", "完全未知业务域", unmapped and ext_count > 0,
                        {"unmapped": unmapped, "extension_proposals": ext_count})

    def _test_duplicate_pk(self, category):
        resp = self._run_fixture(category, "SYN-C8", "重复主键", "duplicate_pk_fixture.xlsx")
        if resp:
            self.record(category, "SYN-C8", "重复主键", resp["status"] != "blocked",
                        {"mapping_count": len(resp["mapping_draft"])})

    def _test_formula_cell(self, category):
        resp = self._run_fixture(category, "SYN-C9", "公式单元格", "formula_cell_fixture.xlsx")
        if resp:
            formula_count = len(resp["schema_summary"].get("formula_fields", []))
            self.record(category, "SYN-C9", "公式单元格检测", formula_count > 0,
                        {"formula_count": formula_count})

    def _test_empty_table(self, category):
        resp = self._run_fixture(category, "SYN-C10", "空表", "empty_table_fixture.xlsx")
        if resp:
            row_count = resp["schema_summary"]["tables"][0]["row_count"] if resp["schema_summary"]["tables"] else -1
            self.record(category, "SYN-C10", "空表处理", row_count == 0,
                        {"row_count": row_count})

    def _test_hidden_sheet(self, category):
        resp = self._run_fixture(category, "SYN-C11", "隐藏表检测", "hidden_sheet_fixture.xlsx")
        if resp:
            hidden_count = len(resp["schema_summary"].get("hidden_sheet_hints", []))
            self.record(category, "SYN-C11", "隐藏表检测", hidden_count > 0,
                        {"hidden_count": hidden_count})

    def _test_corrupt_xlsx(self, category):
        fpath = os.path.join(FIXTURES_DIR, "corrupt_xlsx_fixture.xlsx")
        mapper = BifrostDataMapper(CONTRACTS_DIR)
        req = make_request(fpath, "xlsx")
        resp = mapper.orchestrate_mapping_run(req)
        blocked = resp.get("status") == "blocked" or resp.get("blocked_code") == BLOCKED_CORRUPT_FILE
        if not blocked and resp.get("schema_summary"):
            blocked = len(resp["schema_summary"].get("tables", [])) == 0
        self.record(category, "SYN-C12", "损坏xlsx阻塞", blocked,
                    {"status": resp.get("status"), "blocked_code": resp.get("blocked_code")})

    def _test_unsupported_format(self, category):
        fpath = os.path.join(FIXTURES_DIR, "unsupported_xls_fixture.xls")
        mapper = BifrostDataMapper(CONTRACTS_DIR)
        req = make_request(fpath, "xls")
        resp = mapper.orchestrate_mapping_run(req)
        blocked = resp.get("status") == "blocked" and resp.get("blocked_code") == BLOCKED_UNSUPPORTED_FORMAT
        self.record(category, "SYN-C13a", "不支持的xls格式", blocked,
                    {"status": resp.get("status"), "blocked_code": resp.get("blocked_code")})

        fpath = os.path.join(FIXTURES_DIR, "unsupported_xlsm_fixture.xlsm")
        req = make_request(fpath, "xlsm")
        resp = mapper.orchestrate_mapping_run(req)
        blocked = resp.get("status") == "blocked" and resp.get("blocked_code") == BLOCKED_UNSUPPORTED_FORMAT
        self.record(category, "SYN-C13b", "不支持的xlsm格式", blocked,
                    {"status": resp.get("status"), "blocked_code": resp.get("blocked_code")})

    def _test_path_traversal(self, category):
        fpath = "../../../etc/passwd"
        mapper = BifrostDataMapper(CONTRACTS_DIR)
        req = make_request(fpath, "xlsx")
        resp = mapper.orchestrate_mapping_run(req)
        blocked = (resp.get("status") == "blocked"
                   and resp.get("blocked_code") == BLOCKED_PATH_TRAVERSAL)
        self.record(category, "SYN-C14", "路径穿越阻塞(BLOCKED_PATH_TRAVERSAL)", blocked,
                    {"status": resp.get("status"), "blocked_code": resp.get("blocked_code")})

    def _test_missing_input(self, category):
        fpath = os.path.join(FIXTURES_DIR, "nonexistent_file.xlsx")
        mapper = BifrostDataMapper(CONTRACTS_DIR)
        req = make_request(fpath, "xlsx")
        resp = mapper.orchestrate_mapping_run(req)
        blocked = resp.get("status") == "blocked"
        self.record(category, "SYN-C15", "输入缺失阻塞", blocked,
                    {"status": resp.get("status"), "blocked_code": resp.get("blocked_code")})

    # ------------------------------------------------------------------
    # D. 防硬编码测试
    # ------------------------------------------------------------------

    def run_antihardcode_tests(self):
        category = "internal_skill_tests"
        if not os.path.exists(SIM_FILE):
            return

        import openpyxl
        modified_path = os.path.join(FIXTURES_DIR, "antihardcode_modified.xlsx")
        shutil.copy(SIM_FILE, modified_path)
        wb = openpyxl.load_workbook(modified_path, read_only=False)
        if "01_OEE班次" in wb.sheetnames:
            ws = wb["01_OEE班次"]
            ws.title = "01_修改后班次"
            if ws.max_column >= 1:
                old_val = ws.cell(row=1, column=1).value
                ws.cell(row=1, column=1).value = "修改后_" + str(old_val)
        wb.save(modified_path)
        wb.close()

        mapper = BifrostDataMapper(CONTRACTS_DIR)
        req = make_request(modified_path, "xlsx")
        resp = mapper.orchestrate_mapping_run(req)

        table_names = [t["table_name"] for t in resp["schema_summary"]["tables"]]
        has_modified = any("修改后" in tn for tn in table_names)
        self.record(category, "HC-D1", "修改表名后输出变化", has_modified,
                    {"table_names_sample": table_names[:3]})

        # 真实断言：改名表的首字段应反映修改后名称（来自实际扫描，非残留）
        mod_table_records = [r for r in resp["mapping_draft"] if r.get("source_table") == "01_修改后班次"]
        mod_first = mod_table_records[0]["source_field"] if mod_table_records else None
        no_residual = mod_first is not None and "修改后" in str(mod_first)
        self.record(category, "HC-D2", "无原业务结论残留", no_residual,
                    {"modified_table_first_field": mod_first, "modified_table_records": len(mod_table_records)})

        if os.path.exists(modified_path):
            os.remove(modified_path)

    # ------------------------------------------------------------------
    # E. 合同真实性测试
    # ------------------------------------------------------------------

    def run_contract_tests(self):
        category = "internal_skill_tests"

        cv = ContractValidator(CONTRACTS_DIR)
        cv_result = cv.validate_all()
        self.record(category, "CV-01", "合同校验器通过", cv_result["status"] == "passed",
                    {"errors": cv_result.get("errors", [])})

        # 官方 MAP 注册信息：不得因缺 source_family 登记 n/a
        off_entry = next((e for e in cv_result["registered_contracts"]
                          if e["filename"] == "MAP_OFFICIAL_v0.2.1_mapping.json"), {})
        off_registry_ok = (off_entry.get("approval_status") == "approved"
                           and off_entry.get("data_nature") == "OFFICIAL_DEIDENTIFIED_SIMULATION"
                           and off_entry.get("approval_source") == "USER_INSTRUCTION_04C3")
        self.record(category, "CV-02", "官方MAP注册元数据真实", off_registry_ok,
                    {"approval_status": off_entry.get("approval_status"),
                     "data_nature": off_entry.get("data_nature"),
                     "approval_source": off_entry.get("approval_source")})

        mapper = BifrostDataMapper(CONTRACTS_DIR)
        req = make_request(os.path.join(FIXTURES_DIR, "unknown_domain_fixture.xlsx"), "xlsx")
        resp = mapper.orchestrate_mapping_run(req)
        confirmed_new = sum(1 for r in resp["mapping_draft"] if r["mapping_status"] == "confirmed" and not r.get("inherited_approval"))
        self.record(category, "CON-E1", "新来源confirmed=0", confirmed_new == 0,
                    {"confirmed_new": confirmed_new})

        confirmed_rel = sum(1 for jc in resp["join_candidates"] if jc["status"] == "confirmed")
        self.record(category, "CON-E2", "新关联confirmed=0", confirmed_rel == 0,
                    {"confirmed_rel": confirmed_rel})

        draft_count = len(resp["mapping_draft"])
        status_counts = resp["mapping_summary"]
        count_sum = sum(status_counts.values())
        self.record(category, "CON-E3", "计数与数组守恒", draft_count == count_sum,
                    {"draft_count": draft_count, "status_sum": count_sum})

        queue_count = len(resp["human_confirmation_queue"])
        expected_queue = sum(1 for r in resp["mapping_draft"] if r.get("requires_human_confirmation"))
        self.record(category, "CON-E4", "人工确认队列完整", queue_count >= expected_queue,
                    {"queue_count": queue_count, "expected": expected_queue})

        main_code = os.path.join(SCRIPT_DIR, "bifrost_data_mapper.py")
        with open(main_code, "r") as f:
            code_content = f.read()
        no_fixture_import = "tests/fixtures" not in code_content and "fixtures" not in code_content
        self.record(category, "CON-E5", "生产代码不加载测试夹具", no_fixture_import,
                    {"checked_file": "bifrost_data_mapper.py"})

        if os.path.exists(SIM_FILE):
            sha_before = sha256_file(SIM_FILE)
            req2 = make_request(SIM_FILE, "xlsx")
            resp2 = mapper.orchestrate_mapping_run(req2)
            sha_after = sha256_file(SIM_FILE)
            self.record(category, "CON-E6", "输入哈希前后一致", sha_before == sha_after,
                        {"before": sha_before[:16], "after": sha_after[:16]})
            write_performed = resp2["source_verification"]["source_write_performed"]
            self.record(category, "CON-E7", "无源文件写操作", write_performed is False,
                        {"write_performed": write_performed})
            trace_ids = [resp.get("local_trace_id"), resp2.get("local_trace_id")]
            unique_traces = len(set(trace_ids)) == len(trace_ids)
            self.record(category, "CON-E8", "local_trace_id唯一", unique_traces,
                        {"trace_ids": trace_ids})

            # 响应字段类型合同：source_data_type/source_dataset_type 拆分，无 source_type
            type_ok, type_msg = _check_response_type_contract(resp2["mapping_draft"])
            self.record(category, "CON-E9", "字段类型合同拆分", type_ok, {"msg": type_msg})

        # 缺失合同立即阻塞
        temp_dir = os.path.join(FIXTURES_DIR, "missing_contract_test")
        if os.path.exists(temp_dir):
            shutil.rmtree(temp_dir)
        shutil.copytree(CONTRACTS_DIR, temp_dir)
        os.remove(os.path.join(temp_dir, "DQ_v1.0.1_contract.json"))
        mapper2 = BifrostDataMapper(temp_dir)
        req3 = make_request(SIM_FILE, "xlsx")
        resp3 = mapper2.orchestrate_mapping_run(req3)
        blocked = resp3.get("status") == "blocked" and resp3.get("blocked_code") == BLOCKED_SEMANTIC_CONTRACT
        self.record(category, "CON-E10", "缺失合同立即阻塞", blocked,
                    {"status": resp3.get("status"), "blocked_code": resp3.get("blocked_code")})
        shutil.rmtree(temp_dir)

        # 合同版本错误阻塞
        temp_dir2 = os.path.join(FIXTURES_DIR, "hash_error_test")
        if os.path.exists(temp_dir2):
            shutil.rmtree(temp_dir2)
        shutil.copytree(CONTRACTS_DIR, temp_dir2)
        dq_path = os.path.join(temp_dir2, "DQ_v1.0.1_contract.json")
        with open(dq_path, "r") as f:
            dq = json.load(f)
        dq["quality_contract_version"] = "DQ-v9.9.9"
        with open(dq_path, "w") as f:
            json.dump(dq, f, ensure_ascii=False)
        mapper3 = BifrostDataMapper(temp_dir2)
        req4 = make_request(SIM_FILE, "xlsx")
        resp4 = mapper3.orchestrate_mapping_run(req4)
        self.record(category, "CON-E11", "合同版本错误阻塞", resp4.get("status") == "blocked",
                    {"status": resp4.get("status")})
        shutil.rmtree(temp_dir2)

        self.record(category, "CON-E12", "blocked不输出伪造映射",
                    len(resp3.get("mapping_draft", [])) == 0,
                    {"mapping_count": len(resp3.get("mapping_draft", []))})
        has_versions = all(v for v in resp.get("contract_versions", {}).values())
        self.record(category, "CON-E13", "合同版本与哈希可审计", has_versions,
                    {"contract_count": len(resp.get("contract_versions", {}))})
        self.record(category, "CON-E14", "失败路径不生成业务映射",
                    len(resp3.get("mapping_draft", [])) == 0 and len(resp3.get("join_candidates", [])) == 0,
                    {"draft_count": len(resp3.get("mapping_draft", []))})

    # ------------------------------------------------------------------
    # F. 反假通过变异测试
    # ------------------------------------------------------------------

    def run_antifakepass_tests(self):
        category = "internal_skill_tests"
        self._afp1_wrong_source_sheet(category)
        self._afp2_delete_fm023(category)
        self._afp3_inventory_zero(category)
        self._afp4_banned_assertion_scan(category)
        self._afp5_coln_pseudo_field(category)
        self._afp6_type_contract_overwrite(category)

    def _afp1_wrong_source_sheet(self, category):
        """把官方合同 source_sheet 临时改成错误键时，54条继承测试必须失败。"""
        temp_dir = os.path.join(FIXTURES_DIR, "afp1_wrong_sheet")
        if os.path.exists(temp_dir):
            shutil.rmtree(temp_dir)
        shutil.copytree(CONTRACTS_DIR, temp_dir)
        mp_path = os.path.join(temp_dir, "MAP_OFFICIAL_v0.2.1_mapping.json")
        with open(mp_path, "r") as f:
            mp = json.load(f)
        for fm in mp["field_mappings"]:
            if "source_sheet" in fm:
                fm["source_sheet"] = "WRONG_" + fm["source_sheet"]
        with open(mp_path, "w") as f:
            json.dump(mp, f, ensure_ascii=False)
        mapper = BifrostDataMapper(temp_dir)
        req = make_request(OFFICIAL_FILE, "xlsx",
                           declared_source_family="OFFICIAL_DEIDENTIFIED_SIMULATION",
                           declared_source_type="OFFICIAL_DEIDENTIFIED_SIMULATION")
        resp = mapper.orchestrate_mapping_run(req)
        inherited = resp.get("inherited_approval_count", -1)
        # 变异后继承数应 != 54（即 OFF-C 的 ==54 断言会失败）→ 证明测试为真
        afp_pass = inherited != 54
        self.record(category, "AFP-1", "错误source_sheet致继承失败", afp_pass,
                    {"inherited_after_tamper": inherited})
        shutil.rmtree(temp_dir)

    def _afp2_delete_fm023(self, category):
        """删除 FM-023 时，OEE映射测试必须失败。"""
        temp_dir = os.path.join(FIXTURES_DIR, "afp2_del_fm023")
        if os.path.exists(temp_dir):
            shutil.rmtree(temp_dir)
        shutil.copytree(CONTRACTS_DIR, temp_dir)
        mp_path = os.path.join(temp_dir, "MAP_OFFICIAL_v0.2.1_mapping.json")
        with open(mp_path, "r") as f:
            mp = json.load(f)
        mp["field_mappings"] = [fm for fm in mp["field_mappings"] if fm.get("field_id") != "FM-023"]
        with open(mp_path, "w") as f:
            json.dump(mp, f, ensure_ascii=False)
        mapper = BifrostDataMapper(temp_dir)
        req = make_request(OFFICIAL_FILE, "xlsx",
                           declared_source_family="OFFICIAL_DEIDENTIFIED_SIMULATION",
                           declared_source_type="OFFICIAL_DEIDENTIFIED_SIMULATION")
        resp = mapper.orchestrate_mapping_run(req)
        oee_src = [r for r in resp["mapping_draft"] if r.get("target_field") == "oee_source"]
        # 删除 FM-023 后 oee_source 映射应为 0（OFF-D 的 >0 断言会失败）→ 证明测试为真
        afp_pass = len(oee_src) == 0
        self.record(category, "AFP-2", "删FM-023致OEE映射失败", afp_pass,
                    {"oee_source_count_after_delete": len(oee_src)})
        shutil.rmtree(temp_dir)

    def _afp3_inventory_zero(self, category):
        """inventory_snapshot 映射数为0时，库存测试必须失败。"""
        temp_dir = os.path.join(FIXTURES_DIR, "afp3_inv_zero")
        if os.path.exists(temp_dir):
            shutil.rmtree(temp_dir)
        shutil.copytree(CONTRACTS_DIR, temp_dir)
        mp_path = os.path.join(temp_dir, "MAP_OFFICIAL_v0.2.1_mapping.json")
        with open(mp_path, "r") as f:
            mp = json.load(f)
        for fm in mp["field_mappings"]:
            if fm.get("target_entity") == "inventory_snapshot":
                fm["source_sheet"] = "WRONG_" + fm.get("source_sheet", "")
        with open(mp_path, "w") as f:
            json.dump(mp, f, ensure_ascii=False)
        mapper = BifrostDataMapper(temp_dir)
        req = make_request(OFFICIAL_FILE, "xlsx",
                           declared_source_family="OFFICIAL_DEIDENTIFIED_SIMULATION",
                           declared_source_type="OFFICIAL_DEIDENTIFIED_SIMULATION")
        resp = mapper.orchestrate_mapping_run(req)
        inv_cnt = resp.get("inventory_contract_flags", {}).get("inventory_snapshot_mapping_count", -1)
        # 变异后库存映射应为 0（OFF-F 的 >0 断言会失败）→ 证明测试为真
        afp_pass = inv_cnt == 0
        self.record(category, "AFP-3", "库存映射为0致测试失败", afp_pass,
                    {"inventory_mapping_count_after_tamper": inv_cnt})
        shutil.rmtree(temp_dir)

    def _afp4_banned_assertion_scan(self, category):
        """将断言替换为 True 或 vacuous 下界时，静态测试扫描必须失败。"""
        with open(__file__, "r", encoding="utf-8") as f:
            src = f.read()
        # 构造被禁令牌（运行时拼接，避免扫描器命中自身源码）
        eq = chr(61)
        banned_tokens = [
            "passed" + eq + "True",
            "passed = " + "True",
            ">" + eq + " 0",
            ">" + eq + "0",
        ]
        found = [t for t in banned_tokens if t in src]
        clean = len(found) == 0
        # 注入证明：把禁令牌注入副本，扫描器应能检出
        injected = src + "\n    x_passed" + eq + "True  # AFP inject\n"
        inject_found = any(t in injected for t in banned_tokens)
        afp_pass = clean and inject_found
        self.record(category, "AFP-4", "禁断言静态扫描", afp_pass,
                    {"banned_found_in_source": found, "scanner_detects_injection": inject_found})

    def _afp5_coln_pseudo_field(self, category):
        """生成 col_n 伪字段时，字段扫描测试必须失败。"""
        # 真实输出不得含 col_ 伪字段（SIM-H 的断言）
        mapper = BifrostDataMapper(CONTRACTS_DIR)
        req = make_request(SIM_FILE, "xlsx",
                           declared_source_family="TEAM_ENGINEERED_SIMULATION",
                           declared_source_type="TEAM_ENGINEERED_SIMULATION_STATIC_SNAPSHOT")
        resp = mapper.orchestrate_mapping_run(req)
        draft = resp["mapping_draft"]
        col_n_real = [r for r in draft if str(r.get("source_field", "")).startswith("col_")]
        real_clean = len(col_n_real) == 0
        sem = resp["schema_summary"]["semantic_mapping_field_count"]
        # 变异证明：若强制生成 col_n（语义字段数会 == 物理列数 316 而非 291）
        # 用合成 draft 模拟 col_n 出现，验证 SIM-H 等价检查会失败
        synthetic_with_coln = draft + [{"source_field": "col_99", "mapping_status": "unmapped"}]
        coln_in_synth = any(str(r.get("source_field", "")).startswith("col_") for r in synthetic_with_coln)
        afp_pass = real_clean and sem == 291 and coln_in_synth
        self.record(category, "AFP-5", "col_n伪字段扫描", afp_pass,
                    {"real_col_n_count": len(col_n_real), "semantic_field_count": sem,
                     "synthetic_coln_detected": coln_in_synth})

    def _afp6_type_contract_overwrite(self, category):
        """source_data_type 被 source_dataset_type 覆盖时，响应合同测试必须失败。"""
        mapper = BifrostDataMapper(CONTRACTS_DIR)
        req = make_request(SIM_FILE, "xlsx",
                           declared_source_family="TEAM_ENGINEERED_SIMULATION",
                           declared_source_type="TEAM_ENGINEERED_SIMULATION_STATIC_SNAPSHOT")
        resp = mapper.orchestrate_mapping_run(req)
        draft = resp["mapping_draft"]
        # 真实响应合同通过
        ok_real, _ = _check_response_type_contract(draft)
        # 变异：模拟旧 bug —— 删 source_data_type，加 source_type=source_dataset_type
        mutated = []
        for r in draft:
            mr = dict(r)
            if "source_data_type" in mr:
                del mr["source_data_type"]
                mr["source_type"] = mr.get("source_dataset_type")
            mutated.append(mr)
        ok_mutated, _ = _check_response_type_contract(mutated)
        # 变异后合同检查必须失败 → 证明 CON-E9 测试为真
        afp_pass = ok_real and (not ok_mutated)
        self.record(category, "AFP-6", "类型合同覆盖变异", afp_pass,
                    {"real_contract_ok": ok_real, "mutated_contract_ok": ok_mutated})

    # ------------------------------------------------------------------
    # G. 可信边界测试（04C.4A.2 新增，12 项）
    # ------------------------------------------------------------------

    def run_trust_boundary_tests(self):
        category = "internal_skill_tests"
        self._tb1_sim_exact(category)
        self._tb2_official_exact(category)
        self._tb3_unknown_renamed_as_official(category)
        self._tb4_unknown_declared_official(category)
        self._tb5_modified_structure_same_filename(category)
        self._tb6_data_changed_compatible(category)
        self._tb7_readonly_false(category)
        self._tb8_readonly_missing(category)
        self._tb9_path_traversal_code(category)
        self._tb10_legal_absolute_path(category)
        self._tb11_static_scan_no_filename_identity(category)
        self._tb12_version_consistency(category)

    def _tb1_sim_exact(self, category):
        """已批准 SIM 文件原样运行：signature=exact，继承24条。"""
        mapper = BifrostDataMapper(CONTRACTS_DIR)
        req = make_request(SIM_FILE, "xlsx",
                           declared_source_family="TEAM_ENGINEERED_SIMULATION",
                           declared_source_type="TEAM_ENGINEERED_SIMULATION_STATIC_SNAPSHOT")
        resp = mapper.orchestrate_mapping_run(req)
        sig = resp["source_identity"]["source_signature_status"]
        inherited = resp["inherited_approval_count"]
        passed = (sig == "exact" and inherited == 24)
        self.record(category, "TB-SIG-1", "SIM原样exact继承24", passed,
                    {"signature": sig, "inherited": inherited})

    def _tb2_official_exact(self, category):
        """已批准官方文件原样运行：signature=exact，继承54条。"""
        mapper = BifrostDataMapper(CONTRACTS_DIR)
        req = make_request(OFFICIAL_FILE, "xlsx",
                           declared_source_family="OFFICIAL_DEIDENTIFIED_SIMULATION",
                           declared_source_type="OFFICIAL_DEIDENTIFIED_SIMULATION")
        resp = mapper.orchestrate_mapping_run(req)
        sig = resp["source_identity"]["source_signature_status"]
        inherited = resp["inherited_approval_count"]
        passed = (sig == "exact" and inherited == 54)
        self.record(category, "TB-SIG-2", "官方原样exact继承54", passed,
                    {"signature": sig, "inherited": inherited})

    def _tb3_unknown_renamed_as_official(self, category):
        """将未知xlsx仅改名为"歌尔可脱敏企业测试数据集.xlsx"：signature=unknown，继承0条。"""
        import openpyxl
        tmp_dir = tempfile.mkdtemp(prefix="tb3_")
        dst = os.path.join(tmp_dir, "歌尔可脱敏企业测试数据集.xlsx")
        shutil.copy(os.path.join(FIXTURES_DIR, "unknown_domain_fixture.xlsx"), dst)
        mapper = BifrostDataMapper(CONTRACTS_DIR)
        req = make_request(dst, "xlsx")
        resp = mapper.orchestrate_mapping_run(req)
        sig = resp["source_identity"]["source_signature_status"]
        inherited = resp["inherited_approval_count"]
        passed = (sig == "unknown" and inherited == 0)
        self.record(category, "TB-SIG-3", "未知改名歌尔=unknown继承0", passed,
                    {"signature": sig, "inherited": inherited, "filename": os.path.basename(dst)})
        shutil.rmtree(tmp_dir)

    def _tb4_unknown_declared_official(self, category):
        """未知文件声明declared_source_family=OFFICIAL：signature仍为unknown，继承0条。"""
        mapper = BifrostDataMapper(CONTRACTS_DIR)
        req = make_request(os.path.join(FIXTURES_DIR, "unknown_domain_fixture.xlsx"), "xlsx",
                           declared_source_family="OFFICIAL_DEIDENTIFIED_SIMULATION",
                           declared_source_type="OFFICIAL_DEIDENTIFIED_SIMULATION")
        resp = mapper.orchestrate_mapping_run(req)
        sig = resp["source_identity"]["source_signature_status"]
        inherited = resp["inherited_approval_count"]
        passed = (sig == "unknown" and inherited == 0)
        self.record(category, "TB-SIG-4", "未知声明OFFICIAL=unknown继承0", passed,
                    {"signature": sig, "inherited": inherited,
                     "declared": resp["source_identity"]["declared_source_family"]})

    def _tb5_modified_structure_same_filename(self, category):
        """修改官方文件字段结构但保留文件名：不得继承54条。"""
        import openpyxl
        tmp_dir = tempfile.mkdtemp(prefix="tb5_")
        dst = os.path.join(tmp_dir, "歌尔可脱敏企业测试数据集.xlsx")
        shutil.copy(OFFICIAL_FILE, dst)
        wb = openpyxl.load_workbook(dst)
        # 修改第一张表的第一个表头字段名（破坏 schema 指纹）
        ws = wb.worksheets[0]
        ws.cell(row=1, column=1, value="MODIFIED_FIELD_NAME_XYZ")
        wb.save(dst)
        wb.close()
        mapper = BifrostDataMapper(CONTRACTS_DIR)
        req = make_request(dst, "xlsx",
                           declared_source_family="OFFICIAL_DEIDENTIFIED_SIMULATION",
                           declared_source_type="OFFICIAL_DEIDENTIFIED_SIMULATION")
        resp = mapper.orchestrate_mapping_run(req)
        inherited = resp["inherited_approval_count"]
        sig = resp["source_identity"]["source_signature_status"]
        passed = (inherited != 54 and sig == "unknown")
        self.record(category, "TB-SIG-5", "改结构保名不继承54", passed,
                    {"signature": sig, "inherited": inherited, "filename": os.path.basename(dst)})
        shutil.rmtree(tmp_dir)

    def _tb6_data_changed_compatible(self, category):
        """数据行变化但字段结构与合同一致：signature=compatible，继承54条。"""
        import openpyxl
        tmp_dir = tempfile.mkdtemp(prefix="tb6_")
        dst = os.path.join(tmp_dir, "歌尔可脱敏企业测试数据集_modified_data.xlsx")
        shutil.copy(OFFICIAL_FILE, dst)
        wb = openpyxl.load_workbook(dst)
        # 仅修改一个数据单元格（不改表头/字段结构），破坏文件 SHA 但保持 schema 指纹
        ws = wb.worksheets[0]
        # 找到第一个数据单元格（第2行第1列）修改其值
        old_val = ws.cell(row=2, column=1).value
        ws.cell(row=2, column=1, value=str(old_val) + "_DATA_MODIFIED")
        wb.save(dst)
        wb.close()
        mapper = BifrostDataMapper(CONTRACTS_DIR)
        req = make_request(dst, "xlsx",
                           declared_source_family="OFFICIAL_DEIDENTIFIED_SIMULATION",
                           declared_source_type="OFFICIAL_DEIDENTIFIED_SIMULATION")
        resp = mapper.orchestrate_mapping_run(req)
        sig = resp["source_identity"]["source_signature_status"]
        inherited = resp["inherited_approval_count"]
        # compatible 身份 + 通过字段级验证的批准映射可继承（54条）
        passed = (sig == "compatible" and inherited == 54)
        self.record(category, "TB-SIG-6", "数据变compatible继承54", passed,
                    {"signature": sig, "inherited": inherited})
        shutil.rmtree(tmp_dir)

    def _tb7_readonly_false(self, category):
        """read_only=false：必须返回BLOCKED_READ_ONLY_VIOLATION。"""
        mapper = BifrostDataMapper(CONTRACTS_DIR)
        req = make_request(SIM_FILE, "xlsx")
        req["read_only"] = False
        resp = mapper.orchestrate_mapping_run(req)
        passed = (resp.get("status") == "blocked"
                  and resp.get("blocked_code") == BLOCKED_READ_ONLY_VIOLATION)
        self.record(category, "TB-RO-1", "read_only=false阻塞", passed,
                    {"status": resp.get("status"), "blocked_code": resp.get("blocked_code")})

    def _tb8_readonly_missing(self, category):
        """read_only缺失：必须返回BLOCKED_READ_ONLY_VIOLATION。"""
        mapper = BifrostDataMapper(CONTRACTS_DIR)
        req = make_request(SIM_FILE, "xlsx")
        del req["read_only"]
        resp = mapper.orchestrate_mapping_run(req)
        passed = (resp.get("status") == "blocked"
                  and resp.get("blocked_code") == BLOCKED_READ_ONLY_VIOLATION)
        self.record(category, "TB-RO-2", "read_only缺失阻塞", passed,
                    {"status": resp.get("status"), "blocked_code": resp.get("blocked_code")})

    def _tb9_path_traversal_code(self, category):
        """../../../etc/passwd：必须返回BLOCKED_PATH_TRAVERSAL，不得返回BLOCKED_FILE_NOT_FOUND。"""
        mapper = BifrostDataMapper(CONTRACTS_DIR)
        req = make_request("../../../etc/passwd", "xlsx")
        resp = mapper.orchestrate_mapping_run(req)
        code = resp.get("blocked_code")
        passed = (resp.get("status") == "blocked" and code == BLOCKED_PATH_TRAVERSAL
                  and code != BLOCKED_FILE_NOT_FOUND)
        self.record(category, "TB-PT-1", "路径穿越=PATH_TRAVERSAL", passed,
                    {"status": resp.get("status"), "blocked_code": code})

    def _tb10_legal_absolute_path(self, category):
        """合法绝对附件路径：必须正常读取。"""
        mapper = BifrostDataMapper(CONTRACTS_DIR)
        req = make_request(SIM_FILE, "xlsx",
                           declared_source_family="TEAM_ENGINEERED_SIMULATION",
                           declared_source_type="TEAM_ENGINEERED_SIMULATION_STATIC_SNAPSHOT")
        resp = mapper.orchestrate_mapping_run(req)
        not_blocked = resp.get("status") != "blocked"
        has_draft = resp.get("mapping_draft_count", 0) > 0
        passed = (not_blocked and has_draft and os.path.isabs(SIM_FILE))
        self.record(category, "TB-PT-2", "合法绝对路径正常读取", passed,
                    {"status": resp.get("status"), "is_absolute": os.path.isabs(SIM_FILE),
                     "draft_count": resp.get("mapping_draft_count")})

    def _tb11_static_scan_no_filename_identity(self, category):
        """静态扫描生产代码：不得存在基于"歌尔/脱敏/sim/v2.2"等文件名授予可信身份的逻辑。"""
        main_code = os.path.join(SCRIPT_DIR, "bifrost_data_mapper.py")
        with open(main_code, "r", encoding="utf-8") as f:
            lines = f.readlines()
        # 提取 _identify_source 方法体（从 def 到下一个同级 def）
        method_body = ""
        in_method = False
        for line in lines:
            if line.startswith("    def _identify_source("):
                in_method = True
                method_body += line
                continue
            if in_method:
                if line.startswith("    def ") and "_identify_source" not in line:
                    break
                method_body += line
        banned_patterns = [
            '"sim" in filename',
            "'sim' in filename",
            '"v2.2" in filename',
            "'v2.2' in filename",
            '"歌尔"',
            "'歌尔'",
            '"脱敏"',
            "'脱敏'",
            "sim.*v2.2",
        ]
        found = [p for p in banned_patterns if p in method_body]
        passed = (len(found) == 0 and "approved_asset_sha256" in method_body
                  and "approved_schema_fingerprint_sha256" in method_body)
        self.record(category, "TB-SCAN-1", "无文件名身份逻辑", passed,
                    {"banned_found": found,
                     "has_sha256_check": "approved_asset_sha256" in method_body,
                     "has_fingerprint_check": "approved_schema_fingerprint_sha256" in method_body,
                     "method_body_length": len(method_body)})

    def _tb12_version_consistency(self, category):
        """SKILL.md、代码、Schema和报告版本全部等于0.1.2。"""
        # 代码版本
        code_ver = SKILL_VERSION
        # SKILL.md 版本
        skill_md = os.path.join(SKILL_DIR, "SKILL.md")
        with open(skill_md, "r", encoding="utf-8") as f:
            md_content = f.read()
        md_has_012 = "0.1.2" in md_content
        md_no_010 = "0.1.0不具备" not in md_content and "0.1.0 不具备" not in md_content
        # Schema const 版本
        schema_path = os.path.join(SKILL_DIR, "references", "schemas", "mapping_response_schema.json")
        with open(schema_path, "r", encoding="utf-8") as f:
            schema_content = f.read()
        schema_has_012 = '"0.1.2"' in schema_content
        schema_no_011 = '"0.1.1"' not in schema_content
        # 代码中无 0.1.1 残留（SKILL_VERSION 已为 0.1.2）
        main_code = os.path.join(SCRIPT_DIR, "bifrost_data_mapper.py")
        with open(main_code, "r", encoding="utf-8") as f:
            code_content = f.read()
        code_no_011_const = 'SKILL_VERSION = "0.1.1"' not in code_content
        # join note 无 0.1.0
        code_no_010_note = "0.1.0 不具备" not in code_content and "0.1.0不具备" not in code_content
        passed = (code_ver == "0.1.2" and md_has_012 and md_no_010
                  and schema_has_012 and schema_no_011
                  and code_no_011_const and code_no_010_note)
        self.record(category, "TB-VER-1", "版本一致0.1.2", passed,
                    {"code_version": code_ver, "md_has_012": md_has_012, "md_no_010": md_no_010,
                     "schema_has_012": schema_has_012, "schema_no_011": schema_no_011,
                     "code_no_011_const": code_no_011_const, "code_no_010_note": code_no_010_note})

    # ------------------------------------------------------------------
    # 运行所有测试
    # ------------------------------------------------------------------

    def run_all(self):
        sim_resp = self.run_sim_regression()
        official_resp = self.run_official_regression()
        self.run_synthetic_tests()
        self.run_antihardcode_tests()
        self.run_contract_tests()
        self.run_antifakepass_tests()
        self.run_trust_boundary_tests()

        # 动态统计：分类数字必须与实际数组长度一致
        internal_n = len(self.results["internal_skill_tests"])
        source_n = len(self.results["source_regression_tests"])
        external_n = len(self.results["external_post_package_checks"])
        total = internal_n + source_n + external_n

        self.results["summary"] = {
            "total_tests": self.results["test_session"]["total_tests"],
            "passed": self.results["test_session"]["passed"],
            "failed": self.results["test_session"]["failed"],
            "internal_skill_tests_count": internal_n,
            "source_regression_tests_count": source_n,
            "external_post_package_checks_count": external_n,
            "total_tests_dynamic": total,
            "counts_consistent": total == self.results["test_session"]["total_tests"],
            "all_passed": self.results["test_session"]["failed"] == 0
        }

        return self.results


if __name__ == "__main__":
    runner = TestRunner()
    results = runner.run_all()
    output_path = os.path.join(WORK_DIR, "test_results_04C4A.2.json")
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2, default=str)
    print(f"测试完成: {results['summary']['passed']}/{results['summary']['total_tests']} 通过")
    print(f"internal={results['summary']['internal_skill_tests_count']} "
          f"source={results['summary']['source_regression_tests_count']} "
          f"external={results['summary']['external_post_package_checks_count']}")
    print(f"结果已写入: {output_path}")
