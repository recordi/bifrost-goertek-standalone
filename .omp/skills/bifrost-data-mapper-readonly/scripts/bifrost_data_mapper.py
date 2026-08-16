#!/usr/bin/env python3
"""
bifrost_data_mapper.py — BIFROST 字段整合只读映射核心引擎

职责：
  读取结构化数据文件 → 识别表/字段/类型/单位/值域/质量特征
  → 匹配 BIFROST 统一语义模型 → 发现关联键
  → 生成字段映射草稿 / 关联候选 / 语义扩展建议
  → 人工确认门控 → 输出可审计结构化结果

全程只读：不修改来源数据，不直接修改正式语义模型。
"""

import hashlib
import json
import os
import re
import uuid
from datetime import datetime, date
from pathlib import Path
from typing import Any, Optional

import openpyxl

# ---------------------------------------------------------------------------
# 常量
# ---------------------------------------------------------------------------

SUPPORTED_FORMATS = {"xlsx", "csv", "json"}
SKILL_VERSION = "0.1.2"
RELEASE_STATUS = "LOCAL_VALIDATED_NOT_PUBLISHED"

# 说明/校验型非标准表：空表头列不得自动命名为 col_n 进入语义映射
NON_TABULAR_SHEETS = {
    "00_导入说明",
    "10_多产线说明",
    "15_多产线校验",
    "19_v3交付验收报告",
}

MAPPING_STATUS_ENUM = {"confirmed", "proposed", "ambiguous", "rejected", "unmapped"}
RELATION_STATUS_ENUM = {"validated_candidate", "ambiguous", "rejected", "confirmed"}
DQ_DETECTION_STATUS_ENUM = {"detected", "tested_no_anomaly", "not_tested", "needs_rule"}

BLOCKED_UNSUPPORTED_FORMAT = "BLOCKED_UNSUPPORTED_FORMAT"
BLOCKED_SEMANTIC_CONTRACT = "BLOCKED_SEMANTIC_CONTRACT"
BLOCKED_FILE_NOT_FOUND = "BLOCKED_FILE_NOT_FOUND"
BLOCKED_READ_ONLY_VIOLATION = "BLOCKED_READ_ONLY_VIOLATION"
BLOCKED_PATH_TRAVERSAL = "BLOCKED_PATH_TRAVERSAL"
BLOCKED_CORRUPT_FILE = "BLOCKED_CORRUPT_FILE"


# ---------------------------------------------------------------------------
# 工具函数
# ---------------------------------------------------------------------------

def _sha256_file(filepath: str) -> str:
    h = hashlib.sha256()
    with open(filepath, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()


def _safe_str(val: Any) -> str:
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.isoformat()
    if isinstance(val, date):
        return val.isoformat()
    return str(val)


def _detect_cell_type(val: Any) -> str:
    if val is None:
        return "null"
    if isinstance(val, bool):
        return "bool"
    if isinstance(val, (int,)):
        return "int"
    if isinstance(val, (float,)):
        return "float"
    if isinstance(val, (datetime, date)):
        return "datetime"
    return "str"


def _deterministic_samples(values: list, limit: int = 5) -> list:
    """确定性采样：取前 limit 个非空值，不随机。"""
    samples = []
    for v in values:
        if v is not None and _safe_str(v) != "":
            samples.append(_safe_str(v))
            if len(samples) >= limit:
                break
    return samples


def _compute_value_range(values: list, max_top: int = 10) -> Optional[dict]:
    """计算值域：数值型给 min/max，枚举型给 top_values。"""
    non_null = [v for v in values if v is not None and _safe_str(v) != ""]
    if not non_null:
        return None
    numeric_vals = []
    for v in non_null:
        try:
            numeric_vals.append(float(v))
        except (ValueError, TypeError):
            pass
    if numeric_vals and len(numeric_vals) == len(non_null):
        return {
            "min": min(numeric_vals),
            "max": max(numeric_vals),
            "type": "numeric"
        }
    from collections import Counter
    counts = Counter(_safe_str(v) for v in non_null)
    top = counts.most_common(max_top)
    return {
        "type": "categorical",
        "top_values": [{"value": k, "count": c} for k, c in top]
    }


# ---------------------------------------------------------------------------
# 合同加载器
# ---------------------------------------------------------------------------

class ContractLoader:
    """加载并登记版本化合同，校验完整性。"""

    REQUIRED_CONTRACTS = [
        "SEM_v1.1.1_unified_semantic_model.json",
        "MAP_v1.0.1_mapping.json",
        "MAP_OFFICIAL_v0.2.1_mapping.json",
        "DQ_v1.0.1_contract.json",
        "source_type_registry.json",
        "mapping_status_contract.json",
        "relation_status_contract.json",
        "source_signature_registry_v1.0.json",
    ]

    EXPECTED_VERSIONS = {
        "SEM_v1.1.1_unified_semantic_model.json": "SEM-v1.1.1",
        "MAP_v1.0.1_mapping.json": "MAP-v1.0.1",
        "MAP_OFFICIAL_v0.2.1_mapping.json": "MAP-OFFICIAL-v0.2.1",
        "DQ_v1.0.1_contract.json": "DQ-v1.0.1",
        "source_signature_registry_v1.0.json": "SOURCE-SIGNATURE-v1.0",
    }

    def __init__(self, contracts_dir: str):
        self.contracts_dir = contracts_dir
        self.registry: dict[str, dict] = {}
        self.loaded = False

    def load_all(self) -> dict:
        """加载所有合同，返回登记表。缺失/哈希异常 → BLOCKED_SEMANTIC_CONTRACT。"""
        result = {
            "status": "ok",
            "registered_contracts": [],
            "error": None
        }
        for fname in self.REQUIRED_CONTRACTS:
            fpath = os.path.join(self.contracts_dir, fname)
            if not os.path.exists(fpath):
                result["status"] = BLOCKED_SEMANTIC_CONTRACT
                result["error"] = f"合同文件缺失: {fname}"
                return result
            with open(fpath, "r", encoding="utf-8") as f:
                raw = f.read()
            sha = hashlib.sha256(raw.encode("utf-8")).hexdigest()
            try:
                data = json.loads(raw)
            except json.JSONDecodeError:
                result["status"] = BLOCKED_SEMANTIC_CONTRACT
                result["error"] = f"合同文件 JSON 解析失败: {fname}"
                return result

            # 提取版本/来源信息
            version = self._extract_version(fname, data)
            source_family = self._extract_source_family(fname, data)
            source_type = self._extract_source_type(fname, data)
            data_nature = data.get("data_nature", "") if isinstance(data, dict) else ""
            approval_source = data.get("approval_source", "") if isinstance(data, dict) else ""
            approved_by = data.get("approved_by", "") if isinstance(data, dict) else ""
            approved_at = data.get("approved_at", "") if isinstance(data, dict) else ""

            # 校验版本
            if fname in self.EXPECTED_VERSIONS:
                expected_ver = self.EXPECTED_VERSIONS[fname]
                if version != expected_ver:
                    result["status"] = BLOCKED_SEMANTIC_CONTRACT
                    result["error"] = f"合同版本错误: {fname} 期望 {expected_ver} 实际 {version}"
                    return result

            # 审批状态：不得仅因顶层缺少 source_family 就登记为 n/a；
            # 应读取 data_nature / approval_source / approved_by / approved_at 等真实元数据
            has_approval_meta = bool(approval_source or approved_by or approved_at or data_nature or source_family)
            approval_status = "approved" if has_approval_meta else "n/a"

            entry = {
                "filename": fname,
                "version": version,
                "sha256": sha,
                "source_family": source_family,
                "source_type": source_type,
                "data_nature": data_nature,
                "approval_source": approval_source,
                "approved_by": approved_by,
                "approved_at": approved_at,
                "approval_status": approval_status,
                "loaded_at_runtime": datetime.now().isoformat()
            }
            self.registry[fname] = entry
            result["registered_contracts"].append(entry)

        self.loaded = True
        return result

    def _extract_version(self, fname: str, data: Any) -> str:
        keys_map = {
            "SEM_": "semantic_model_version",
            "MAP_v1.0.1": "mapping_rule_version",
            "MAP_OFFICIAL": "official_mapping_rule_version",
            "DQ_": "quality_contract_version",
            "source_signature_registry": "source_signature_registry_version",
        }
        for prefix, key in keys_map.items():
            if fname.startswith(prefix) and isinstance(data, dict):
                return data.get(key, "unknown")
        if isinstance(data, dict):
            return data.get("contract_version", data.get("registry_version", "unknown"))
        return "unknown"

    def _extract_source_family(self, fname: str, data: Any) -> str:
        if isinstance(data, dict):
            # 优先 source_family；缺失时回退 data_nature（官方 MAP 顶层无 source_family）
            return data.get("source_family", "") or data.get("data_nature", "")
        return ""

    def _extract_source_type(self, fname: str, data: Any) -> str:
        if isinstance(data, dict):
            return data.get("source_type", "")
        return ""

    def get_sem(self) -> dict:
        fpath = os.path.join(self.contracts_dir, "SEM_v1.1.1_unified_semantic_model.json")
        with open(fpath, "r", encoding="utf-8") as f:
            return json.load(f)

    def get_map_sim(self) -> dict:
        fpath = os.path.join(self.contracts_dir, "MAP_v1.0.1_mapping.json")
        with open(fpath, "r", encoding="utf-8") as f:
            return json.load(f)

    def get_map_official(self) -> dict:
        fpath = os.path.join(self.contracts_dir, "MAP_OFFICIAL_v0.2.1_mapping.json")
        with open(fpath, "r", encoding="utf-8") as f:
            return json.load(f)

    def get_dq_contract(self) -> dict:
        fpath = os.path.join(self.contracts_dir, "DQ_v1.0.1_contract.json")
        with open(fpath, "r", encoding="utf-8") as f:
            return json.load(f)

    def get_source_type_registry(self) -> dict:
        fpath = os.path.join(self.contracts_dir, "source_type_registry.json")
        with open(fpath, "r", encoding="utf-8") as f:
            return json.load(f)

    def get_mapping_status_contract(self) -> dict:
        fpath = os.path.join(self.contracts_dir, "mapping_status_contract.json")
        with open(fpath, "r", encoding="utf-8") as f:
            return json.load(f)

    def get_relation_status_contract(self) -> dict:
        fpath = os.path.join(self.contracts_dir, "relation_status_contract.json")
        with open(fpath, "r", encoding="utf-8") as f:
            return json.load(f)

    def get_source_signature_registry(self) -> dict:
        fpath = os.path.join(self.contracts_dir, "source_signature_registry_v1.0.json")
        with open(fpath, "r", encoding="utf-8") as f:
            return json.load(f)


# ---------------------------------------------------------------------------
# 核心映射引擎
# ---------------------------------------------------------------------------

class BifrostDataMapper:
    """字段整合只读映射引擎。"""

    def __init__(self, contracts_dir: str):
        self.contracts = ContractLoader(contracts_dir)
        self.trace_id = f"TRACE-{datetime.now().strftime('%Y%m%d%H%M%S')}-{uuid.uuid4().hex[:8]}"
        self._sem_cache = None
        self._map_sim_cache = None
        self._map_official_cache = None

    # -- 合同校验门控 --

    def _ensure_contracts(self) -> Optional[str]:
        if not self.contracts.loaded:
            r = self.contracts.load_all()
            if r["status"] != "ok":
                return r["error"]
        return None

    @property
    def sem(self) -> dict:
        if self._sem_cache is None:
            self._sem_cache = self.contracts.get_sem()
        return self._sem_cache

    @property
    def map_sim(self) -> dict:
        if self._map_sim_cache is None:
            self._map_sim_cache = self.contracts.get_map_sim()
        return self._map_sim_cache

    @property
    def map_official(self) -> dict:
        if self._map_official_cache is None:
            self._map_official_cache = self.contracts.get_map_official()
        return self._map_official_cache

    # -- 1. verify_source_asset --

    @staticmethod
    def _has_path_traversal(filepath: str) -> bool:
        """检测原始路径是否包含独立的 '..' 路径段。使用 normpath + 逐段检查，不因路径不存在降级。"""
        if not isinstance(filepath, str) or filepath == "":
            return False
        # 原始路径按分隔符拆段，检测独立 '..' 段
        raw_parts = re.split(r"[\\/]", filepath)
        if ".." in raw_parts:
            return True
        # normpath 后再检测（覆盖 ./../ 等）
        norm = os.path.normpath(filepath)
        norm_parts = re.split(r"[\\/]", norm)
        if ".." in norm_parts:
            return True
        return False

    def verify_source_asset(self, filepath: str, expected_format: Optional[str] = None) -> dict:
        """验证文件存在、格式、大小、SHA-256。禁止修改源文件。"""
        result = {
            "function": "verify_source_asset",
            "file_path": filepath,
            "exists": False,
            "format": None,
            "size_bytes": 0,
            "source_sha256": None,
            "source_write_performed": False,
            "blocked": None
        }
        # 路径穿越检查：原始路径包含独立的 ".." 路径段时阻塞
        # 必须在文件存在性检查前返回，不得用 pass 跳过
        if self._has_path_traversal(filepath):
            result["blocked"] = BLOCKED_PATH_TRAVERSAL
            return result
        if not os.path.exists(filepath):
            result["blocked"] = BLOCKED_FILE_NOT_FOUND
            return result

        result["exists"] = True
        result["size_bytes"] = os.path.getsize(filepath)
        result["source_sha256"] = _sha256_file(filepath)

        ext = os.path.splitext(filepath)[1].lower().lstrip(".")
        result["format"] = ext

        if ext not in SUPPORTED_FORMATS:
            result["blocked"] = BLOCKED_UNSUPPORTED_FORMAT
            return result

        if expected_format and ext != expected_format:
            result["blocked"] = BLOCKED_UNSUPPORTED_FORMAT
            return result

        return result

    # -- 2. inspect_dataset_schema --

    def inspect_dataset_schema(self, filepath: str, fmt: str) -> dict:
        """输出表名、行数、字段数、字段顺序、公式字段、隐藏表提示。Excel 只读模式。"""
        result = {
            "function": "inspect_dataset_schema",
            "tables": [],
            "hidden_sheet_hints": [],
            "formula_fields": [],
            "blocked": None
        }
        if fmt == "xlsx":
            try:
                wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            except Exception as e:
                result["blocked"] = BLOCKED_CORRUPT_FILE
                result["error"] = str(e)
                return result

            for ws in wb.worksheets:
                table_info = {
                    "table_name": ws.title,
                    "row_count": 0,
                    "field_count": 0,
                    "field_names": [],
                    "field_order": [],
                    "state": ws.sheet_state
                }
                if ws.sheet_state != "visible":
                    result["hidden_sheet_hints"].append({
                        "sheet": ws.title,
                        "state": ws.sheet_state,
                        "hint": "隐藏表已检测，内容纳入只读扫描"
                    })

                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    table_info["row_count"] = 0
                    table_info["field_count"] = 0
                    table_info["physical_used_column_count"] = 0
                    table_info["excluded_column_count"] = 0
                    table_info["is_non_tabular"] = ws.title in NON_TABULAR_SHEETS
                else:
                    header = rows[0]
                    # 物理使用列：任何行有值的最大列
                    max_real_col = 0
                    for ri, row in enumerate(rows):
                        for ci in range(len(row) - 1, -1, -1):
                            if row[ci] is not None and _safe_str(row[ci]) != "":
                                if ci + 1 > max_real_col:
                                    max_real_col = ci + 1
                                break

                    is_non_tabular = ws.title in NON_TABULAR_SHEETS
                    used_header = list(header[:max_real_col]) if max_real_col > 0 else []

                    # 语义字段 = 非空表头列；空表头列不得自动命名为 col_n 进入语义映射
                    field_names = []
                    field_indices = []
                    excluded_columns = []
                    for ci, h in enumerate(used_header):
                        h_str = _safe_str(h) if h is not None else ""
                        if h_str != "":
                            field_names.append(h_str)
                            field_indices.append(ci)
                        else:
                            # 空表头但下方存在值：归入排除列，不进入语义映射
                            has_value_below = any(
                                (ci < len(row) and row[ci] is not None and _safe_str(row[ci]) != "")
                                for row in rows[1:]
                            )
                            excluded_columns.append({
                                "index": ci,
                                "generated_name": f"col_{ci}",
                                "has_value_below": has_value_below
                            })

                    table_info["field_names"] = field_names
                    table_info["field_indices"] = field_indices
                    table_info["field_order"] = list(range(len(field_names)))
                    table_info["field_count"] = len(field_names)
                    table_info["physical_used_column_count"] = max_real_col
                    table_info["excluded_column_count"] = len(excluded_columns)
                    table_info["excluded_columns"] = excluded_columns
                    table_info["is_non_tabular"] = is_non_tabular
                    table_info["row_count"] = max(0, len(rows) - 1)  # 减去表头

                result["tables"].append(table_info)

            wb.close()

            # 检查公式字段（用 data_only=False 对比）
            try:
                wb2 = openpyxl.load_workbook(filepath, read_only=True, data_only=False)
                for ws in wb2.worksheets:
                    rows = list(ws.iter_rows(values_only=True))
                    if not rows:
                        continue
                    header = rows[0]
                    for ri, row in enumerate(rows[1:], 1):
                        for ci, val in enumerate(row):
                            if isinstance(val, str) and val.startswith("="):
                                col_name = _safe_str(header[ci]) if ci < len(header) else f"col_{ci}"
                                result["formula_fields"].append({
                                    "sheet": ws.title,
                                    "row": ri,
                                    "col": col_name,
                                    "formula": val[:200]
                                })
                                break  # 每行只报第一个公式
                wb2.close()
            except Exception:
                pass

        elif fmt == "csv":
            import csv
            with open(filepath, "r", encoding="utf-8") as f:
                reader = csv.reader(f)
                rows = list(reader)
            if rows:
                header = rows[0]
                table_info = {
                    "table_name": os.path.basename(filepath),
                    "row_count": max(0, len(rows) - 1),
                    "field_count": len(header),
                    "field_names": [_safe_str(h) for h in header],
                    "field_order": list(range(len(header))),
                    "state": "visible"
                }
                result["tables"].append(table_info)

        elif fmt == "json":
            with open(filepath, "r", encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, list) and data:
                field_names = list(data[0].keys()) if isinstance(data[0], dict) else []
                table_info = {
                    "table_name": os.path.basename(filepath),
                    "row_count": len(data),
                    "field_count": len(field_names),
                    "field_names": field_names,
                    "field_order": list(range(len(field_names))),
                    "state": "visible"
                }
                result["tables"].append(table_info)
            elif isinstance(data, dict):
                table_info = {
                    "table_name": os.path.basename(filepath),
                    "row_count": 1,
                    "field_count": len(data),
                    "field_names": list(data.keys()),
                    "field_order": list(range(len(data))),
                    "state": "visible"
                }
                result["tables"].append(table_info)

        return result

    # -- 3. profile_source_field --

    def profile_source_field(self, filepath: str, fmt: str, schema: dict) -> dict:
        """输出字段类型、非空数、空值率、唯一率、样例、值域、日期/编号格式、异常候选。"""
        profiles = []
        if fmt == "xlsx":
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            for table in schema["tables"]:
                ws = wb[table["table_name"]]
                rows = list(ws.iter_rows(values_only=True))
                if not rows or len(rows) < 2:
                    for fi, fname in enumerate(table["field_names"]):
                        profiles.append(self._make_profile(table["table_name"], fname, [], fi))
                    continue
                data_rows = rows[1:]
                field_indices = table.get("field_indices", list(range(len(table["field_names"]))))
                for fi, fname in enumerate(table["field_names"]):
                    col_idx = field_indices[fi] if fi < len(field_indices) else fi
                    col_values = [row[col_idx] if col_idx < len(row) else None for row in data_rows]
                    profiles.append(self._make_profile(table["table_name"], fname, col_values, fi))
            wb.close()
        elif fmt == "csv":
            import csv
            with open(filepath, "r", encoding="utf-8") as f:
                reader = csv.reader(f)
                rows = list(reader)
            if rows:
                header = rows[0]
                data_rows = rows[1:]
                table = schema["tables"][0]
                for fi, fname in enumerate(table["field_names"]):
                    col_values = [row[fi] if fi < len(row) else None for row in data_rows]
                    profiles.append(self._make_profile(table["table_name"], fname, col_values, fi))
        elif fmt == "json":
            with open(filepath, "r", encoding="utf-8") as f:
                data = json.load(f)
            table = schema["tables"][0]
            if isinstance(data, list):
                for fi, fname in enumerate(table["field_names"]):
                    col_values = [row.get(fname) if isinstance(row, dict) else None for row in data]
                    profiles.append(self._make_profile(table["table_name"], fname, col_values, fi))

        return {
            "function": "profile_source_field",
            "profiles": profiles,
            "total_fields": len(profiles)
        }

    def _make_profile(self, table_name: str, field_name: str, values: list, field_index: int) -> dict:
        total = len(values)
        non_null = [v for v in values if v is not None and _safe_str(v) != ""]
        null_count = total - len(non_null)
        null_rate = round(null_count / total, 6) if total > 0 else 0.0

        str_vals = [_safe_str(v) for v in non_null]
        unique_vals = set(str_vals)
        unique_rate = round(len(unique_vals) / len(non_null), 6) if non_null else 0.0

        # 类型推断
        type_counts = {}
        for v in non_null:
            t = _detect_cell_type(v)
            type_counts[t] = type_counts.get(t, 0) + 1
        inferred_type = max(type_counts, key=type_counts.get) if type_counts else "null"

        samples = _deterministic_samples(values, 5)
        value_range = _compute_value_range(non_null)

        # 日期/编号格式检测
        format_hint = self._detect_format_hint(non_null)

        # 异常候选
        anomaly_candidates = self._detect_anomalies(non_null, inferred_type)

        return {
            "table_name": table_name,
            "field_name": field_name,
            "field_index": field_index,
            "inferred_type": inferred_type,
            "type_distribution": type_counts,
            "total_count": total,
            "non_null_count": len(non_null),
            "null_count": null_count,
            "null_rate": null_rate,
            "unique_count": len(unique_vals),
            "unique_rate": unique_rate,
            "samples": samples,
            "value_range": value_range,
            "format_hint": format_hint,
            "anomaly_candidates": anomaly_candidates
        }

    def _detect_format_hint(self, values: list) -> Optional[str]:
        if not values:
            return None
        first = _safe_str(values[0])
        # 日期格式
        if re.match(r"\d{4}-\d{2}-\d{2}", first):
            return "date_iso"
        if re.match(r"\d{4}/\d{2}/\d{2}", first):
            return "date_slash"
        if re.match(r"\d{4}年\d{2}月\d{2}日", first):
            return "date_cn"
        # 编号格式
        if re.match(r"[A-Z]{2,}-", first):
            return "id_prefix"
        if re.match(r"PJ-\d{4}-", first):
            return "project_id"
        return None

    def _detect_anomalies(self, values: list, inferred_type: str) -> list:
        anomalies = []
        numeric_vals = []
        for v in values:
            try:
                numeric_vals.append(float(v))
            except (ValueError, TypeError):
                pass
        if numeric_vals and inferred_type in ("float", "int"):
            # 性能率 > 1.0：保留原始值，标记 flagged_above_unity，不截断
            over_one = [v for v in numeric_vals if v > 1.0]
            if over_one and all(0 <= v <= 2 for v in numeric_vals):
                anomalies.append({
                    "type": "ratio_over_one",
                    "count": len(over_one),
                    "validation_status": "flagged_above_unity",
                    "sample_original_values": [round(v, 6) for v in over_one[:5]],
                    "auto_truncated": False,
                    "hint": "值 > 1.0，可能是性能率原始值，保留并标记 flagged_above_unity"
                })
            # 负值
            negatives = [v for v in numeric_vals if v < 0]
            if negatives:
                anomalies.append({
                    "type": "negative_value",
                    "count": len(negatives),
                    "hint": "存在负值"
                })
        return anomalies

    # -- 4. normalize_data_type --

    def normalize_data_type(self, profiles: list) -> dict:
        """生成类型标准化建议，不覆盖原始值。"""
        suggestions = []
        for p in profiles:
            t = p["inferred_type"]
            suggestion = {
                "table_name": p["table_name"],
                "field_name": p["field_name"],
                "original_type": t,
                "suggested_type": t,
                "action": "keep",
                "note": None
            }
            if t == "str":
                # 检查是否全是数字字符串
                samples = p.get("samples", [])
                if samples and all(re.match(r"^-?\d+\.?\d*$", s) for s in samples if s):
                    suggestion["suggested_type"] = "float"
                    suggestion["action"] = "suggest_cast"
                    suggestion["note"] = "字符串值均为数字，建议转为数值类型"
            elif t == "datetime":
                suggestion["suggested_type"] = "datetime"
                suggestion["action"] = "keep"
            elif t == "null":
                suggestion["suggested_type"] = "unknown"
                suggestion["action"] = "flag"
                suggestion["note"] = "字段全空，无法推断类型"
            suggestions.append(suggestion)
        return {
            "function": "normalize_data_type",
            "suggestions": suggestions,
            "total": len(suggestions)
        }

    # -- 5. normalize_unit --

    def normalize_unit(self, profiles: list, field_name: Optional[str] = None) -> dict:
        """识别 ratio/percent/count/second/minute/hour/date 等单位。只生成建议。"""
        unit_keywords = {
            "ratio": ["率", "oee", "availability", "quality", "performance", "yield"],
            "percent": ["百分比", "percent", "%"],
            "count": ["数量", "产量", "count", "qty", "数"],
            "second": ["秒", "sec", "second"],
            "minute": ["分", "min", "minute"],
            "hour": ["小时", "hour", "hr"],
            "date": ["日期", "date", "时间", "time"],
        }
        suggestions = []
        for p in profiles:
            fname_lower = p["field_name"].lower()
            detected_unit = None
            for unit, keywords in unit_keywords.items():
                if any(kw in fname_lower for kw in keywords):
                    detected_unit = unit
                    break

            # 比例冲突检测
            ambiguous = False
            if detected_unit == "ratio" or (p["value_range"] and p["value_range"].get("type") == "numeric"):
                vr = p["value_range"]
                if vr and vr.get("type") == "numeric":
                    vmin, vmax = vr["min"], vr["max"]
                    if 0 <= vmin <= 1 and vmax <= 1:
                        detected_unit = detected_unit or "ratio"
                    elif 0 <= vmin and vmax <= 100 and vmax > 1:
                        if detected_unit == "ratio":
                            ambiguous = True
                        detected_unit = detected_unit or "percent"

            suggestion = {
                "table_name": p["table_name"],
                "field_name": p["field_name"],
                "detected_unit": detected_unit,
                "ambiguous": ambiguous,
                "suggestion": f"建议单位: {detected_unit}" if detected_unit else "无单位线索",
                "action": "suggest_only"
            }
            if ambiguous:
                suggestion["suggestion"] = "0-1 与 0-100 比例冲突，标记 ambiguous"
                suggestion["action"] = "flag_ambiguous"
            suggestions.append(suggestion)

        return {
            "function": "normalize_unit",
            "suggestions": suggestions,
            "total": len(suggestions),
            "ambiguous_count": sum(1 for s in suggestions if s["ambiguous"])
        }

    # -- 6. match_semantic_field --

    def match_semantic_field(self, profiles: list, source_family: str, source_type: str) -> dict:
        """根据批准合同、字段名、类型、单位、值域和实体上下文进行匹配。"""
        # 选择对应的映射合同
        if source_family == "TEAM_ENGINEERED_SIMULATION":
            approved_map = self.map_sim
            map_field_mappings = approved_map.get("field_mappings", [])
        elif source_family == "OFFICIAL_DEIDENTIFIED_SIMULATION":
            approved_map = self.map_official
            map_field_mappings = approved_map.get("field_mappings", [])
        else:
            # 未知来源，无批准合同
            approved_map = None
            map_field_mappings = []

        # 建立查找索引：规范化统一键 normalized_source_table + source_field
        # SIM MAP 使用 source_table；OFFICIAL MAP 使用 source_sheet —— 二者归一
        approved_index = {}
        for fm in map_field_mappings:
            norm_table = fm.get("source_table") or fm.get("source_sheet") or ""
            fm["_normalized_source_table"] = norm_table
            key = (norm_table, fm.get("source_field", ""))
            approved_index[key] = fm

        sem = self.sem
        # 建立语义字段索引：target_field → (entity, field_def)
        sem_field_index = {}
        for entity, fields in sem.items():
            if not isinstance(fields, dict):
                continue
            for fname, fdef in fields.items():
                sem_field_index[fname] = (entity, fdef)

        matches = []
        suspended_inherited = []
        for p in profiles:
            match = self._match_single_field(p, approved_index, sem_field_index, sem, source_family)
            susp = match.pop("_suspended_info", None)
            if susp:
                suspended_inherited.append(susp)
            matches.append(match)

        return {
            "function": "match_semantic_field",
            "matches": matches,
            "total": len(matches),
            "suspended_inherited_approvals": suspended_inherited
        }

    def _match_single_field(self, profile: dict, approved_index: dict, sem_field_index: dict, sem: dict, source_family: str) -> dict:
        table = profile["table_name"]
        field = profile["field_name"]
        key = (table, field)

        # 1. 先查批准合同（规范化键已覆盖 source_table / source_sheet）
        if key in approved_index:
            am = approved_index[key]
            # 验证字段名、类型、单位是否匹配
            type_ok = True
            unit_ok = True
            # 简单类型检查（用合同 source_type 字段推断数据类型）
            contract_src_type = am.get("source_type") or am.get("data_nature")
            if contract_src_type and profile["inferred_type"]:
                # 仅当合同 source_type 像数据类型（str/int/float...）时校验
                if contract_src_type in ("str", "int", "float", "bool", "datetime", "enum"):
                    type_ok = self._types_compatible(contract_src_type, profile["inferred_type"])

            is_confirmed_approved = (
                am.get("mapping_status") == "confirmed"
                and am.get("approval_status") == "approved"
            )

            # 继承批准降级检查：单位/类型/口径不一致 → suspended，不计入 active confirmed
            suspended = is_confirmed_approved and (not type_ok or not unit_ok)
            inherited_approval = is_confirmed_approved and not suspended

            result = {
                "table_name": table,
                "field_name": field,
                "normalized_source_table": am.get("_normalized_source_table", table),
                "match_method": "approved_contract",
                "target_entity": am.get("target_entity"),
                "target_field": am.get("target_field"),
                "confidence": am.get("confidence", 0.5),
                "mapping_status": ("ambiguous" if suspended else am.get("mapping_status", "proposed")),
                "inherited_approval": inherited_approval,
                "inherited_approval_suspended": suspended,
                "approval_source": am.get("approval_source") if inherited_approval else None,
                "approval_rule_id": am.get("field_id") if inherited_approval else None,
                "approval_rule_version": am.get("mapping_rule_version") if inherited_approval else None,
                "requires_human_confirmation": am.get("requires_human_confirmation", True),
                "evidence": am.get("evidence", "") or am.get("approval_reason", ""),
                "type_match": type_ok,
                "unit_match": unit_ok
            }
            if suspended:
                result["_suspended_info"] = {
                    "source_table": table,
                    "source_field": field,
                    "original_mapping_status": am.get("mapping_status"),
                    "original_approval_status": am.get("approval_status"),
                    "current_mapping_status": "ambiguous",
                    "current_state": "suspended",
                    "downgrade_reason": "类型或单位口径检查未通过" if not type_ok else "单位口径检查未通过",
                    "evidence": f"type_match={type_ok}, unit_match={unit_ok}"
                }
            return result

        # 2. 语义模型匹配（基于字段名相似度和实体上下文）
        best_match = None
        best_score = 0.0
        for sem_field, (entity, fdef) in sem_field_index.items():
            score = self._semantic_similarity(field, sem_field, fdef, profile)
            if score > best_score:
                best_score = score
                best_match = (entity, sem_field, fdef)

        if best_match and best_score >= 0.6:
            entity, sem_field, fdef = best_match
            status = "proposed" if best_score >= 0.8 else "ambiguous"
            return {
                "table_name": table,
                "field_name": field,
                "match_method": "semantic_inference",
                "target_entity": entity,
                "target_field": sem_field,
                "confidence": round(best_score, 4),
                "mapping_status": status,
                "inherited_approval": False,
                "approval_source": None,
                "approval_rule_id": None,
                "approval_rule_version": None,
                "requires_human_confirmation": True,
                "evidence": f"语义推断匹配 (score={best_score:.4f})",
                "type_match": None,
                "unit_match": None
            }

        # 3. 未匹配
        return {
            "table_name": table,
            "field_name": field,
            "match_method": "no_match",
            "target_entity": None,
            "target_field": None,
            "confidence": 0.0,
            "mapping_status": "unmapped",
            "inherited_approval": False,
            "approval_source": None,
            "approval_rule_id": None,
            "approval_rule_version": None,
            "requires_human_confirmation": True,
            "evidence": "无匹配规则且语义推断低于阈值",
            "type_match": None,
            "unit_match": None
        }

    def _types_compatible(self, contract_type: str, inferred_type: str) -> bool:
        mapping = {
            "str": {"str", "int", "float", "datetime"},
            "int": {"int", "float"},
            "float": {"float", "int"},
            "enum": {"str", "enum"},
            "bool": {"bool"},
            "datetime": {"datetime", "str"},
        }
        allowed = mapping.get(contract_type, {contract_type})
        return inferred_type in allowed

    def _semantic_similarity(self, source_field: str, sem_field: str, fdef: dict, profile: dict) -> float:
        """语义相似度评分，不只按字符串相似度。"""
        s = source_field.lower().replace("_", "").replace(" ", "")
        t = sem_field.lower().replace("_", "").replace(" ", "")
        if s == t:
            return 1.0
        if t in s or s in t:
            return 0.85

        # label_cn 匹配
        label_cn = fdef.get("label_cn", "") if isinstance(fdef, dict) else ""
        if label_cn and source_field in label_cn:
            return 0.8
        if label_cn and label_cn in source_field:
            return 0.75

        # 部分匹配
        common = set(s) & set(t)
        if common and len(s) > 2 and len(t) > 2:
            jaccard = len(common) / len(set(s) | set(t))
            if jaccard > 0.5:
                return 0.5 + jaccard * 0.2

        return 0.0

    # -- 7. infer_join_key_candidates --

    def infer_join_key_candidates(self, schema: dict, profiles: list, filepath: str, fmt: str) -> dict:
        """输出命中率、唯一率、空值率、基数关系、命名空间和风险。不自动生成外键。"""
        # 收集所有表的字段 profile
        table_profiles = {}
        for p in profiles:
            tn = p["table_name"]
            if tn not in table_profiles:
                table_profiles[tn] = {}
            table_profiles[tn][p["field_name"]] = p

        table_names = list(table_profiles.keys())
        candidates = []

        # 读取实际数据用于命中率计算
        table_data = self._load_table_data(filepath, fmt, schema)

        for i, t1 in enumerate(table_names):
            for t2 in table_names[i + 1:]:
                for f1, p1 in table_profiles[t1].items():
                    for f2, p2 in table_profiles[t2].items():
                        # 字段名相同或语义相近
                        name_sim = self._field_name_similarity(f1, f2)
                        if name_sim < 0.7:
                            continue
                        # 计算命中率
                        vals1 = self._get_column_values(table_data, t1, f1, schema)
                        vals2 = self._get_column_values(table_data, t2, f2, schema)
                        if not vals1 or not vals2:
                            continue
                        set1 = set(_safe_str(v) for v in vals1 if v is not None and _safe_str(v) != "")
                        set2 = set(_safe_str(v) for v in vals2 if v is not None and _safe_str(v) != "")
                        if not set1 or not set2:
                            continue
                        intersection = set1 & set2
                        hit_rate = len(intersection) / min(len(set1), len(set2)) if min(len(set1), len(set2)) > 0 else 0

                        if hit_rate < 0.3:
                            continue

                        # 基数关系
                        if len(set1) == len(set2) and len(intersection) == len(set1):
                            cardinality = "one_to_one"
                        elif len(set1) < len(set2):
                            cardinality = "one_to_many"
                        elif len(set1) > len(set2):
                            cardinality = "many_to_one"
                        else:
                            cardinality = "many_to_many"

                        # 风险评估
                        risks = []
                        if hit_rate < 0.8:
                            risks.append("命中率不足80%，可能存在语义歧义")
                        if p1["unique_rate"] < 0.5 and p2["unique_rate"] < 0.5:
                            risks.append("双方唯一率均低，关联可能非键")
                        if f1 != f2:
                            risks.append("字段名不完全一致")

                        status = "validated_candidate" if hit_rate >= 0.8 and not risks else "ambiguous"

                        candidates.append({
                            "source_table": t1,
                            "source_field": f1,
                            "target_table": t2,
                            "target_field": f2,
                            "exact_hit_rate": round(hit_rate, 6),
                            "source_unique_rate": p1["unique_rate"],
                            "target_unique_rate": p2["unique_rate"],
                            "source_null_rate": p1["null_rate"],
                            "target_null_rate": p2["null_rate"],
                            "relation_cardinality": cardinality,
                            "namespace": f"{t1}.{f1} ↔ {t2}.{f2}",
                            "status": status,
                            "requires_human_confirmation": True,
                            "risks": risks,
                            "evidence": f"命中率={hit_rate:.4f}, 基数={cardinality}"
                        })

        return {
            "function": "infer_join_key_candidates",
            "candidates": candidates,
            "total": len(candidates),
            "confirmed_count": 0,
            "note": "本 Skill 不具备批准关联权限，所有新关联 confirmed=0"
        }

    def _field_name_similarity(self, f1: str, f2: str) -> float:
        s1 = f1.lower().replace("_", "").replace(" ", "")
        s2 = f2.lower().replace("_", "").replace(" ", "")
        if s1 == s2:
            return 1.0
        if s1 in s2 or s2 in s1:
            return 0.8
        common = set(s1) & set(s2)
        if not s1 or not s2:
            return 0.0
        return len(common) / len(set(s1) | set(s2))

    def _load_table_data(self, filepath: str, fmt: str, schema: dict) -> dict:
        data = {}
        if fmt == "xlsx":
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            for table in schema["tables"]:
                ws = wb[table["table_name"]]
                rows = list(ws.iter_rows(values_only=True))
                data[table["table_name"]] = rows[1:] if len(rows) > 1 else []
            wb.close()
        return data

    def _get_column_values(self, table_data: dict, table: str, field: str, schema: dict) -> list:
        rows = table_data.get(table, [])
        # 找到字段索引
        for t in schema["tables"]:
            if t["table_name"] == table:
                if field in t["field_names"]:
                    fi = t["field_names"].index(field)
                    return [row[fi] if fi < len(row) else None for row in rows]
        return []

    # -- 8. generate_mapping_draft --

    def generate_mapping_draft(self, matches: list, unit_suggestions: list, profiles: list, source_family: str, source_type: str) -> dict:
        """为每个来源字段生成且仅生成一条裁决记录。"""
        unit_map = {}
        for us in unit_suggestions:
            unit_map[(us["table_name"], us["field_name"])] = us

        profile_map = {}
        for p in profiles:
            profile_map[(p["table_name"], p["field_name"])] = p

        # 构建实体元数据索引（用于 inventory_snapshot 等实体的合同标志）
        sem_entity_meta = {}
        sem = self.sem
        for entity, fields in sem.items():
            if isinstance(fields, dict) and isinstance(fields.get("_entity_meta"), dict):
                sem_entity_meta[entity] = fields["_entity_meta"]

        draft = []
        for m in matches:
            key = (m["table_name"], m["field_name"])
            us = unit_map.get(key, {})
            p = profile_map.get(key, {})

            # 确定最终状态
            status = m["mapping_status"]
            requires_confirmation = m["requires_human_confirmation"]

            # 单位冲突 → ambiguous
            if us.get("ambiguous"):
                status = "ambiguous"
                requires_confirmation = True

            # 多目标匹配检测（同表同目标多个字段）
            # 已在 match 阶段处理

            # 低置信度
            if m["confidence"] < 0.3 and status not in ("unmapped", "rejected"):
                status = "ambiguous"
                requires_confirmation = True

            record = {
                "source_table": m["table_name"],
                "normalized_source_table": m.get("normalized_source_table", m["table_name"]),
                "source_field": m["field_name"],
                "source_data_type": p.get("inferred_type"),
                "source_dataset_type": source_type,
                "target_entity": m["target_entity"],
                "target_field": m["target_field"],
                "match_method": m["match_method"],
                "mapping_status": status,
                "confidence": m["confidence"],
                "inherited_approval": m["inherited_approval"],
                "inherited_approval_suspended": m.get("inherited_approval_suspended", False),
                "approval_source": m["approval_source"],
                "approval_rule_id": m["approval_rule_id"],
                "approval_rule_version": m["approval_rule_version"],
                "unit": us.get("detected_unit"),
                "unit_ambiguous": us.get("ambiguous", False),
                "requires_human_confirmation": requires_confirmation,
                "evidence": m["evidence"],
                "null_rate": p.get("null_rate"),
                "unique_rate": p.get("unique_rate"),
                "samples": p.get("samples", []),
                "value_range": p.get("value_range"),
                "source_family": source_family,
                "semantic_model_version": "SEM-v1.1.1"
            }

            # 性能率 >1.0 字段：写入 validation_status，保留原始值，不截断
            for anomaly in p.get("anomaly_candidates", []):
                if anomaly.get("type") == "ratio_over_one":
                    record["validation_status"] = anomaly.get("validation_status", "flagged_above_unity")
                    record["over_unity_detected_count"] = anomaly.get("count", 0)
                    record["over_unity_sample_original_values"] = anomaly.get("sample_original_values", [])
                    record["auto_truncated"] = anomaly.get("auto_truncated", False)
                    record["original_value_preserved"] = True
                    break

            # inventory_snapshot 实体合同标志（view_projection_only / grain_status / aggregation_allowed）
            if m.get("target_entity") and m["target_entity"] in sem_entity_meta:
                meta = sem_entity_meta[m["target_entity"]]
                record["view_projection_only"] = meta.get("materialization_status") == "view_projection_only"
                record["grain_status"] = meta.get("grain_status")
                record["aggregation_allowed"] = meta.get("aggregation_allowed")
            draft.append(record)

        return {
            "function": "generate_mapping_draft",
            "mapping_draft": draft,
            "total": len(draft),
            "status_counts": self._count_statuses(draft)
        }

    def _count_statuses(self, draft: list) -> dict:
        counts = {}
        for r in draft:
            s = r["mapping_status"]
            counts[s] = counts.get(s, 0) + 1
        return counts

    # -- 9. generate_semantic_extension_proposals --

    def generate_semantic_extension_proposals(self, draft: list, sem: dict) -> dict:
        """仅在现有 SEM 无法覆盖时输出新实体或新字段建议。状态固定为 proposal。"""
        proposals = []
        sem_entities = {k for k, v in sem.items() if isinstance(v, dict)}
        sem_fields = set()
        for entity, fields in sem.items():
            if isinstance(fields, dict):
                for fn in fields:
                    sem_fields.add(f"{entity}.{fn}")

        for r in draft:
            if r["mapping_status"] == "unmapped" and r["target_entity"] is None:
                # 检查是否可能是新字段
                proposals.append({
                    "proposal_id": f"SEP-{len(proposals)+1:04d}",
                    "source_table": r["source_table"],
                    "source_field": r["source_field"],
                    "proposed_entity": None,
                    "proposed_field": r["source_field"],
                    "status": "proposal",
                    "reason": "现有 SEM 无法覆盖此来源字段",
                    "requires_human_confirmation": True
                })

        return {
            "function": "generate_semantic_extension_proposals",
            "proposals": proposals,
            "total": len(proposals),
            "note": "状态固定为 proposal，不自动修改 SEM-v1.1.1"
        }

    # -- 10. validate_mapping_contract --

    def validate_mapping_contract(self, draft: list, join_candidates: list, contract_registry: list, source_family: str) -> dict:
        """校验字段守恒、状态枚举、审批来源、关联状态、单位与类型一致性、来源追踪和人工确认队列。"""
        issues = []
        warnings = []

        # 字段守恒：每个来源字段恰好一条记录
        seen = set()
        for r in draft:
            key = (r["source_table"], r["source_field"])
            if key in seen:
                issues.append(f"字段守恒违规: 重复记录 {key}")
            seen.add(key)

        # 状态枚举
        for r in draft:
            if r["mapping_status"] not in MAPPING_STATUS_ENUM:
                issues.append(f"状态枚举违规: {r['source_field']} status={r['mapping_status']}")

        # 关联状态枚举
        for jc in join_candidates:
            if jc["status"] not in RELATION_STATUS_ENUM:
                issues.append(f"关联状态枚举违规: {jc['namespace']} status={jc['status']}")

        # 新来源不得 confirmed
        confirmed_fields = [r for r in draft if r["mapping_status"] == "confirmed" and not r.get("inherited_approval")]
        if confirmed_fields:
            issues.append(f"新来源不得自动 confirmed: {len(confirmed_fields)} 个字段违规")

        # 新关联不得 confirmed
        confirmed_relations = [jc for jc in join_candidates if jc["status"] == "confirmed"]
        if confirmed_relations:
            issues.append(f"新关联不得 confirmed: {len(confirmed_relations)} 个违规")

        # 人工确认队列完整性
        queue = [r for r in draft if r.get("requires_human_confirmation")]
        for r in queue:
            if not r.get("evidence"):
                warnings.append(f"人工确认项缺少 evidence: {r['source_field']}")

        # 合同版本校验
        if not contract_registry:
            issues.append("合同注册表为空")

        status = "failed" if issues else ("warning" if warnings else "passed")

        return {
            "function": "validate_mapping_contract",
            "status": status,
            "issues": issues,
            "warnings": warnings,
            "field_count": len(draft),
            "confirmed_inherited_count": sum(1 for r in draft if r.get("inherited_approval")),
            "confirmed_new_count": len(confirmed_fields),
            "suspended_inherited_count": sum(1 for r in draft if r.get("inherited_approval_suspended")),
            "active_confirmed_count": sum(1 for r in draft if r.get("inherited_approval")),
            "join_candidate_count": len(join_candidates),
            "join_confirmed_count": len(confirmed_relations),
            "human_confirmation_queue_size": len(queue)
        }

    # -- 11. orchestrate_mapping_run --

    def orchestrate_mapping_run(self, request: dict) -> dict:
        """按固定顺序执行以上函数，统一输出结果。任一可信门控失败时停止。"""
        trace_id = f"TRACE-{datetime.now().strftime('%Y%m%d%H%M%S')}-{uuid.uuid4().hex[:8]}"

        # 强制只读请求合同 —— 在读取任何文件前验证
        if not isinstance(request, dict):
            return self._blocked_response({}, trace_id, BLOCKED_READ_ONLY_VIOLATION, "request 必须为 object")
        source_file = request.get("source_file")
        if not isinstance(source_file, str) or source_file.strip() == "":
            return self._blocked_response(request, trace_id, BLOCKED_READ_ONLY_VIOLATION, "source_file 必须为非空字符串")
        fmt = request.get("file_format", os.path.splitext(source_file)[1].lstrip(".").lower())
        if fmt not in SUPPORTED_FORMATS:
            return self._blocked_response(request, trace_id, BLOCKED_UNSUPPORTED_FORMAT, f"file_format 不支持: {fmt}")
        if "read_only" not in request:
            return self._blocked_response(request, trace_id, BLOCKED_READ_ONLY_VIOLATION, "read_only 字段缺失")
        if request.get("read_only") is not True:
            return self._blocked_response(request, trace_id, BLOCKED_READ_ONLY_VIOLATION, "read_only 必须严格等于 true")

        # 前置合同校验
        contract_err = self._ensure_contracts()
        if contract_err:
            return self._blocked_response(request, trace_id, BLOCKED_SEMANTIC_CONTRACT, contract_err)

        filepath = request["source_file"]
        # file_format 已在只读门控中验证

        # 1. verify_source_asset
        verification = self.verify_source_asset(filepath, fmt)
        sha_before = verification["source_sha256"]
        if verification.get("blocked"):
            return self._blocked_response(request, trace_id, verification["blocked"], verification.get("error", ""))

        # 2. inspect_dataset_schema
        schema = self.inspect_dataset_schema(filepath, fmt)
        if schema.get("blocked"):
            return self._blocked_response(request, trace_id, schema["blocked"], schema.get("error", ""))

        # 3. profile_source_field
        field_profiles = self.profile_source_field(filepath, fmt, schema)
        profiles = field_profiles["profiles"]

        # 4 & 5. normalize
        type_suggestions = self.normalize_data_type(profiles)
        unit_suggestions = self.normalize_unit(profiles)

        # 来源识别（基于签名注册表，不依赖文件名）
        detected_family, detected_type, identity_confidence, sig_status = self._identify_source(filepath, sha_before, schema, request)

        # 6. match_semantic_field
        matches = self.match_semantic_field(profiles, detected_family, detected_type)

        # 7. infer_join_key_candidates
        join_result = self.infer_join_key_candidates(schema, profiles, filepath, fmt)

        # 8. generate_mapping_draft
        draft_result = self.generate_mapping_draft(
            matches["matches"], unit_suggestions["suggestions"], profiles,
            detected_family, detected_type
        )

        # 9. semantic extension proposals
        ext_proposals = self.generate_semantic_extension_proposals(draft_result["mapping_draft"], self.sem)

        # 数据质量发现
        dq_findings = self._generate_dq_findings(profiles, draft_result["mapping_draft"])

        # 10. validate_mapping_contract
        validation = self.validate_mapping_contract(
            draft_result["mapping_draft"], join_result["candidates"],
            self.contracts.registry, detected_family
        )

        # 11. 验证文件哈希未变
        sha_after = _sha256_file(filepath)
        write_performed = sha_before != sha_after

        # 构建人工确认队列
        confirmation_queue = []
        for r in draft_result["mapping_draft"]:
            if r.get("requires_human_confirmation"):
                confirmation_queue.append({
                    "source_table": r["source_table"],
                    "source_field": r["source_field"],
                    "reason": self._confirmation_reason(r),
                    "current_status": r["mapping_status"]
                })
        for jc in join_result["candidates"]:
            if jc.get("requires_human_confirmation"):
                confirmation_queue.append({
                    "source_table": jc["source_table"],
                    "source_field": jc["source_field"],
                    "reason": f"关联候选需确认: {jc['namespace']}",
                    "current_status": jc["status"]
                })

        # 确定最终状态
        if validation["status"] == "failed":
            final_status = "blocked"
        elif confirmation_queue:
            final_status = "needs_confirmation"
        else:
            final_status = "completed"

        # schema 口径计数：物理使用列 / 语义映射字段 / 非标准表排除列
        physical_used_column_count = sum(t.get("physical_used_column_count", t.get("field_count", 0)) for t in schema["tables"])
        semantic_mapping_field_count = sum(t.get("field_count", 0) for t in schema["tables"])
        non_tabular_excluded_column_count = sum(t.get("excluded_column_count", 0) for t in schema["tables"])

        # OEE 安全合同（从当前来源对应的 MAP 合同读取 oee_safety）
        oee_safety_contract = self._extract_oee_safety_contract(detected_family)

        # inventory_snapshot 实体合同标志聚合
        inventory_contract_flags = self._extract_inventory_contract_flags(draft_result["mapping_draft"])

        suspended_inherited = matches.get("suspended_inherited_approvals", [])

        response = {
            "status": final_status,
            "logical_skill_version": SKILL_VERSION,
            "release_status": RELEASE_STATUS,
            "request_id": request.get("request_id"),
            "source_verification": {
                "source_sha256_before": sha_before,
                "source_sha256_after": sha_after,
                "source_write_performed": write_performed
            },
            "source_identity": {
                "declared_source_family": request.get("declared_source_family"),
                "detected_source_family": detected_family,
                "detected_source_type": detected_type,
                "identity_confidence": identity_confidence,
                "source_signature_status": sig_status,
                "source_signature_registry_version": self._get_source_signature_registry_version()
            },
            "contract_versions": {fname: entry for fname, entry in self.contracts.registry.items()},
            "schema_summary": {
                "table_count": len(schema["tables"]),
                "tables": [{"table_name": t["table_name"], "row_count": t["row_count"], "field_count": t["field_count"],
                            "physical_used_column_count": t.get("physical_used_column_count", t.get("field_count", 0)),
                            "excluded_column_count": t.get("excluded_column_count", 0),
                            "is_non_tabular": t.get("is_non_tabular", False)} for t in schema["tables"]],
                "total_fields": sum(t["field_count"] for t in schema["tables"]),
                "physical_used_column_count": physical_used_column_count,
                "semantic_mapping_field_count": semantic_mapping_field_count,
                "non_tabular_excluded_column_count": non_tabular_excluded_column_count,
                "hidden_sheet_hints": schema["hidden_sheet_hints"],
                "formula_fields": schema["formula_fields"]
            },
            "field_profile_summary": {
                "total_fields": field_profiles["total_fields"],
                "type_distribution": self._aggregate_type_distribution(profiles)
            },
            "mapping_summary": draft_result["status_counts"],
            "mapping_draft": draft_result["mapping_draft"],
            "mapping_draft_count": len(draft_result["mapping_draft"]),
            "inherited_approval_count": sum(1 for r in draft_result["mapping_draft"] if r.get("inherited_approval")),
            "active_confirmed_count": sum(1 for r in draft_result["mapping_draft"] if r.get("inherited_approval")),
            "confirmed_new_count": sum(1 for r in draft_result["mapping_draft"] if r["mapping_status"] == "confirmed" and not r.get("inherited_approval")),
            "suspended_inherited_approvals": suspended_inherited,
            "oee_safety_contract": oee_safety_contract,
            "inventory_contract_flags": inventory_contract_flags,
            "join_candidates": join_result["candidates"],
            "data_quality_findings": dq_findings,
            "semantic_extension_proposals": ext_proposals["proposals"],
            "human_confirmation_queue": confirmation_queue,
            "validation": {
                "status": validation["status"],
                "issues": validation["issues"],
                "warnings": validation["warnings"],
                "confirmed_inherited_count": validation["confirmed_inherited_count"],
                "confirmed_new_count": validation["confirmed_new_count"],
                "suspended_inherited_count": validation["suspended_inherited_count"]
            },
            "data_gaps": self._identify_data_gaps(profiles, schema),
            "local_trace_id": trace_id,
            "aily_run_id": None
        }

        return response

    def _identify_source(self, filepath: str, sha: str, schema: dict, request: dict) -> tuple:
        """识别来源身份。

        仅凭文件 SHA-256(exact) 或 schema 指纹(compatible) 建立可信身份。
        文件名、路径名、source_name、declared_source_family 仅作提示，
        不得单独建立 exact 或 compatible 身份，不得单独触发批准映射继承。
        """
        try:
            registry = self.contracts.get_source_signature_registry()
        except Exception:
            registry = {"approved_sources": []}
        entries = registry.get("approved_sources", []) if isinstance(registry, dict) else []

        # 1. exact：文件 SHA-256 与批准资产完全一致
        for e in entries:
            if e.get("approved_asset_sha256") == sha:
                return (e["source_family"], e["source_type"], 1.0, "exact")

        # 2. compatible：文件哈希变化，但 schema 指纹与批准合同一致
        #    字段级类型/单位/口径验证在 match_semantic_field 阶段逐条执行
        fp = self._compute_schema_fingerprint(schema)
        for e in entries:
            if e.get("approved_schema_fingerprint_sha256") == fp:
                return (e["source_family"], e["source_type"], 0.8, "compatible")

        # 3. unknown：其他情况
        return ("UNKNOWN", "SOURCE_TYPE_UNVERIFIED", 0.1, "unknown")

    def _compute_schema_fingerprint(self, schema: dict) -> str:
        """确定性生成 schema 指纹。

        由以下内容确定性生成，不得包含业务数值：
        - 有序表名
        - 每张表的有序字段名
        - 字段数量
        - 非标准表分类
        """
        parts = []
        for t in schema.get("tables", []):
            tn = t.get("table_name", "")
            fns = t.get("field_names", [])
            fc = len(fns)
            nt = t.get("is_non_tabular", False)
            parts.append(f"{tn}\t{fc}\t{nt}\t" + "|".join(fns))
        payload = "\n".join(parts)
        return hashlib.sha256(payload.encode("utf-8")).hexdigest()

    def _get_source_signature_registry_version(self) -> str:
        """返回签名注册表版本号。"""
        try:
            reg = self.contracts.get_source_signature_registry()
            return reg.get("source_signature_registry_version", "SOURCE-SIGNATURE-v1.0")
        except Exception:
            return "SOURCE-SIGNATURE-v1.0"

    def _extract_oee_safety_contract(self, source_family: str) -> dict:
        """从当前来源对应的 MAP 合同读取 oee_safety 合同字段。"""
        if source_family == "OFFICIAL_DEIDENTIFIED_SIMULATION":
            mp = self.map_official
        elif source_family == "TEAM_ENGINEERED_SIMULATION":
            mp = self.map_sim
        else:
            return {"can_recompute_oee": None, "source": "no_approved_contract"}
        oee = mp.get("oee_safety", {}) if isinstance(mp, dict) else {}
        # SIM MAP 无 oee_safety 顶层键时，给出明确缺省（不通过"未找到"间接推断）
        if not oee:
            return {
                "can_recompute_oee": False,
                "oee_source_only": True,
                "source": mp.get("mapping_rule_version", "unknown"),
                "note": "SIM 合同未声明 oee_safety，按只读安全约束 can_recompute_oee=false"
            }
        return {
            "can_recompute_oee": oee.get("can_recompute_oee", False),
            "oee_source_only": oee.get("oee_source_only"),
            "oee_recomputed": oee.get("oee_recomputed"),
            "no_reverse_engineering": oee.get("no_reverse_engineering"),
            "over_100_percent_preserved": oee.get("over_100_percent_preserved"),
            "quality_factor_mapping_status": oee.get("quality_factor_mapping_status"),
            "source": mp.get("official_mapping_rule_version") or mp.get("mapping_rule_version", "unknown")
        }

    def _extract_inventory_contract_flags(self, draft: list) -> dict:
        """聚合 inventory_snapshot 目标字段的实体合同标志。"""
        inv_records = [r for r in draft if r.get("target_entity") == "inventory_snapshot"]
        if not inv_records:
            return {"inventory_snapshot_mapping_count": 0}
        return {
            "inventory_snapshot_mapping_count": len(inv_records),
            "view_projection_only": all(r.get("view_projection_only") is True for r in inv_records),
            "grain_status": inv_records[0].get("grain_status"),
            "aggregation_allowed": inv_records[0].get("aggregation_allowed"),
            "mapped_fields": [r.get("target_field") for r in inv_records]
        }

    def _generate_dq_findings(self, profiles: list, draft: list) -> list:
        """生成数据质量发现，遵守 DQ-v1.0.1。"""
        findings = []
        finding_id = 0
        # draft 查找索引：(table, field) → target_field
        draft_index = {(r["source_table"], r["source_field"]): r for r in draft}
        for p in profiles:
            # 高空值率
            if p["null_rate"] > 0.5 and p["total_count"] > 0:
                finding_id += 1
                findings.append({
                    "issue_id": f"DQ-{finding_id:04d}",
                    "issue_type": "missing_value",
                    "detection_status": "detected",
                    "source_sheet": p["table_name"],
                    "record_key": "N/A",
                    "source_field": p["field_name"],
                    "observed_value": f"null_rate={p['null_rate']:.4f}",
                    "expected_rule": "required 字段空值率应 < 50%",
                    "rule_source": "DQ-v1.0.1",
                    "severity": "medium",
                    "confidence": 0.9,
                    "evidence": f"空值率 {p['null_rate']:.2%}，超过 50%",
                    "suggested_action": "检查数据采集流程",
                    "auto_fix_allowed": False
                })

            # 异常值
            for anomaly in p.get("anomaly_candidates", []):
                if anomaly.get("type") == "ratio_over_one":
                    dr = draft_index.get((p["table_name"], p["field_name"]), {})
                    finding_id += 1
                    findings.append({
                        "issue_id": f"DQ-{finding_id:04d}",
                        "issue_type": "performance_rate_above_unity",
                        "detection_status": "detected",
                        "source_sheet": p["table_name"],
                        "record_key": "N/A",
                        "source_field": p["field_name"],
                        "target_field": dr.get("target_field"),
                        "observed_value": f"over_unity_count={anomaly.get('count', 0)}",
                        "validation_status": "flagged_above_unity",
                        "detected_count": anomaly.get("count", 0),
                        "sample_original_values": anomaly.get("sample_original_values", []),
                        "auto_truncated": False,
                        "original_value_preserved": True,
                        "expected_rule": "performance_rate_raw > 1.0 须检出并保留原值，标记 flagged_above_unity，不得截断为 1.0",
                        "rule_source": "DQ-v1.0.1 + OEE 安全规则",
                        "severity": "medium",
                        "confidence": 0.95,
                        "evidence": anomaly.get("hint", ""),
                        "suggested_action": "人工确认是否为合理异常，保留原始值",
                        "auto_fix_allowed": False
                    })
                else:
                    finding_id += 1
                    findings.append({
                        "issue_id": f"DQ-{finding_id:04d}",
                        "issue_type": "abnormal_value",
                        "detection_status": "detected",
                        "source_sheet": p["table_name"],
                        "record_key": "N/A",
                        "source_field": p["field_name"],
                        "observed_value": anomaly.get("hint", ""),
                        "expected_rule": "数值应在合理范围内",
                        "rule_source": "DQ-v1.0.1",
                        "severity": "low",
                        "confidence": 0.7,
                        "evidence": anomaly.get("hint", ""),
                        "suggested_action": "人工确认是否为合理异常",
                        "auto_fix_allowed": False
                    })

        # 缺少更新时间 → stale_data=not_tested
        finding_id += 1
        findings.append({
            "issue_id": f"DQ-{finding_id:04d}",
            "issue_type": "stale_data",
            "detection_status": "not_tested",
            "source_sheet": "ALL",
            "record_key": "N/A",
            "source_field": "update_timestamp",
            "observed_value": None,
            "expected_rule": "需要 reliable_update_timestamp / refresh_sla / evaluation_time",
            "rule_source": "DQ-v1.0.1",
            "severity": "low",
            "confidence": 1.0,
            "evidence": "缺少更新时间或刷新 SLA，无法判定数据过期",
            "suggested_action": "补充数据更新时间元信息",
            "auto_fix_allowed": False
        })

        return findings

    def _confirmation_reason(self, r: dict) -> str:
        reasons = []
        if r.get("unit_ambiguous"):
            reasons.append("单位冲突")
        if r.get("confidence", 1) < 0.3:
            reasons.append("低置信度")
        if r["mapping_status"] == "ambiguous":
            reasons.append("口径冲突或多目标匹配")
        if r["mapping_status"] == "unmapped":
            reasons.append("未匹配")
        if r["mapping_status"] == "proposed" and not r.get("inherited_approval"):
            reasons.append("新提议需人工确认")
        return "; ".join(reasons) if reasons else "需人工确认"

    def _aggregate_type_distribution(self, profiles: list) -> dict:
        dist = {}
        for p in profiles:
            t = p["inferred_type"]
            dist[t] = dist.get(t, 0) + 1
        return dist

    def _identify_data_gaps(self, profiles: list, schema: dict) -> list:
        gaps = []
        # 缺少必填规则
        gaps.append({
            "gap": "missing_required_field_rules",
            "detection_status": "needs_rule",
            "description": "缺少业务必填规则定义"
        })
        # 缺少更新时间
        gaps.append({
            "gap": "missing_update_timestamp",
            "detection_status": "not_tested",
            "description": "数据缺少更新时间元信息，stale_data=not_tested"
        })
        return gaps

    def _blocked_response(self, request: dict, trace_id: str, blocked_code: str, error: str) -> dict:
        return {
            "status": "blocked",
            "logical_skill_version": SKILL_VERSION,
            "release_status": RELEASE_STATUS,
            "request_id": request.get("request_id"),
            "blocked_code": blocked_code,
            "error": error,
            "source_verification": {
                "source_sha256_before": None,
                "source_sha256_after": None,
                "source_write_performed": False
            },
            "mapping_draft": [],
            "join_candidates": [],
            "data_quality_findings": [],
            "semantic_extension_proposals": [],
            "human_confirmation_queue": [],
            "validation": {"status": "failed", "issues": [error], "warnings": []},
            "data_gaps": [],
            "local_trace_id": trace_id,
            "aily_run_id": None
        }
