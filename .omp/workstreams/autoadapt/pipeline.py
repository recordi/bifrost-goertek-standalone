"""Read-only raw data -> BIFROST semantic adaptation pipeline.

This module deliberately does not calculate authoritative business KPIs.  It
profiles the source, delegates semantic mapping to the existing mapper Skill,
derives a capability manifest, and returns safe stubs for the next compiler
stage.  Ambiguous mappings and missing inputs remain visible.
"""

from __future__ import annotations

import csv
import hashlib
import json
import re
import subprocess
import sys
import tempfile
from pathlib import Path
from typing import Any


REQUIRED_FIELDS = {
    "oee": {"availability", "performance_rate", "quality_rate"},
    "yield": {"total_output", "good_output"},
    "spc": {"spc_measurement_points", "usl", "lsl", "sample_rule"},
    "mtbf": {"equipment_id", "failure_start", "failure_end"},
    "supply_risk": {"material_id", "required_quantity", "available_quantity"},
}

# A local, deterministic last-resort mapper.  These aliases are intentionally
# conservative: they only propose fields whose names are already present in a
# source profile and never infer values or rows.  Every proposal is marked
# ``fallback_requires_confirmation`` so the confirmation layer cannot silently
# approve it (including exact-alias matches).
DETERMINISTIC_ALIASES = {
    "availability": {"availability", "available_rate", "availability_rate", "开动率"},
    "performance_rate_raw": {"performance_rate", "performance_rate_raw", "performance", "性能率"},
    "quality_rate": {"quality_rate", "quality", "yield", "yield_rate", "良率", "良率(%)", "质量率"},
    "total_output": {"total_output", "actual_output", "actual_qty", "production_qty", "总产量", "实际产量"},
    "good_output": {"good_output", "good_qty", "qualified_qty", "良品数", "良品数量"},
    "oee_source": {"oee", "oee_source", "oee(%)", "oee原值", "oee复算"},
    "line_id": {"line_id", "lineid", "production_line", "line", "产线", "产线id"},
    "line_name": {"line_name", "产线名称"},
    "shift_date": {"shift_date", "production_date", "date", "班次日期", "生产日期", "日期"},
    # These are optional evidence dimensions.  They are proposed when present
    # but never required for a source to enter the KPI path.
    "shift_id": {"shift_id", "shift", "shift_name", "班次", "班次号"},
    "work_order_id": {"work_order_id", "work_order", "order_no", "工单号", "工单"},
    "product_id": {"product_id", "product_code", "model", "product", "产品编码", "产品型号"},
    "equipment_id": {"equipment_id", "machine_id", "equipment", "设备编号", "设备id"},
    "material_id": {"material_id", "material_code", "business_key", "物料编码", "物料号"},
    "defect_type": {"defect_type", "defect_code", "defect", "不良类型", "缺陷类型"},
    "stop_reason": {"stop_reason", "reason_code", "downtime_reason", "停机原因", "停机类别"},
    "process_step": {"process_step", "operation", "process", "工序", "工艺"},
    "downtime_minutes": {"downtime_minutes", "duration_minutes", "stop_minutes", "停机时长", "停机分钟"},
    "defect_count": {"defect_count", "defect_qty", "bad_qty", "defect_total", "不良数", "不良数量"},
    "phase": {"phase", "stage", "阶段"},
}

# Unicode aliases are kept separately so the fallback mapper remains correct
# even when this source file is opened under a legacy Windows code page.
DETERMINISTIC_ALIASES.update({
    "line_id": {"LineID", "\u4ea7\u7ebf", "\u4ea7\u7ebfID"},
    "line_name": {"\u4ea7\u7ebf\u540d\u79f0"},
    "defect_type": {"\u4e0d\u826f\u7c7b\u578b", "\u4e3b\u8981\u4e0d\u826f\u7c7b\u578b", "\u7f3a\u9677\u7c7b\u578b"},
    "defect_count": {"\u4e0d\u826f\u6570", "\u4e0d\u826f\u6570\u91cf", "\u4e0d\u826f\u54c1\u6570", "\u4e0d\u826f\u54c1\u6570\u91cf"},
    "stop_reason": {"\u505c\u673a\u539f\u56e0", "\u6e90\u505c\u673a\u539f\u56e0", "\u6e90\u505c\u673a\u7ec4", "\u505c\u673a\u7c7b\u522b", "\u505c\u673a\u7ec4"},
    "downtime_minutes": {"\u505c\u673a\u65f6\u957f", "\u505c\u673a\u65f6\u957f_\u5206\u949f", "\u6301\u7eed\u65f6\u95f4", "\u505c\u673a\u5206\u949f"},
})


def _normalize_alias(value: Any) -> str:
    """Normalize labels only; do not coerce source cell values."""
    text = str(value or "").strip().casefold()
    return re.sub(r"[\s_\-]+", "", text)


def _deterministic_mapping_draft(profile: dict[str, Any]) -> dict[str, Any]:
    """Build a review-only schema/alias draft when the provider is unavailable.

    The result deliberately contains no canonical rows, measurements, or KPI
    values.  It is a list of candidate field mappings for a human to confirm.
    """
    aliases = {_normalize_alias(alias): target for target, names in DETERMINISTIC_ALIASES.items() for alias in names}
    # Prefer a stable order independent of workbook internals.
    draft: list[dict[str, Any]] = []
    for table in sorted(profile.get("tables", []), key=lambda item: str(item.get("name", ""))):
        table_name = str(table.get("name", ""))
        fields = sorted({str(field) for field in table.get("fields", []) if field is not None}, key=lambda value: (_normalize_alias(value), value))
        for source_field in fields:
            target = aliases.get(_normalize_alias(source_field))
            if not target:
                continue
            percent_hint = "%" in source_field or "百分" in source_field
            draft.append({
                "source_table": table_name,
                "source_field": source_field,
                "target_entity": "production_observation",
                "target_field": target,
                "confidence": 0.70,
                "mapping_status": "proposed",
                "mapping_source": "deterministic_schema_alias_fallback",
                "fallback_requires_confirmation": True,
                "value_mode_hint": "percentage_0_to_100" if percent_hint else "ratio_0_to_1" if target in {"availability", "performance_rate_raw", "quality_rate", "oee_source"} else None,
                "unit_ambiguity": bool(percent_hint and target in {"quality_rate", "oee_source"}),
            })
    return {
        "status": "needs_confirmation",
        "source_identity": {"detected_source_family": "LOCAL_SCHEMA_ALIAS_FALLBACK", "source_sha256": profile.get("source_sha256")},
        "mapping_summary": {"candidate_count": len(draft), "provider": "local_deterministic_alias", "requires_confirmation": True},
        "mapping_draft": draft,
        "data_gaps": ["mapper_no_response", "deterministic_fallback_requires_confirmation"],
        "fallback_used": True,
    }


def _format_for(path: Path) -> str:
    suffix = path.suffix.lower().lstrip(".")
    if suffix not in {"xlsx", "csv", "json"}:
        raise ValueError(f"unsupported_format:{suffix or 'unknown'}")
    return suffix


def _safe_id(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "-", str(value)).strip("-")
    return cleaned[:80] or "SOURCE-LOCAL-001"


def _profile(path: Path) -> dict[str, Any]:
    """Return a bounded source profile without mutating the source."""
    fmt = _format_for(path)
    profile: dict[str, Any] = {
        "file_name": path.name,
        "format": fmt,
        "size_bytes": path.stat().st_size,
        "source_sha256": hashlib.sha256(path.read_bytes()).hexdigest(),
        "tables": [],
    }
    if fmt == "csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            rows = []
            for index, row in enumerate(reader):
                if index >= 100:
                    break
                rows.append(row)
            profile["tables"].append({"name": path.stem, "row_sample": len(rows), "fields": list(reader.fieldnames or [])})
    elif fmt == "json":
        value = json.loads(path.read_text(encoding="utf-8"))
        rows = value if isinstance(value, list) else value.get("rows", []) if isinstance(value, dict) else []
        fields = sorted({key for row in rows[:100] if isinstance(row, dict) for key in row})
        profile["tables"].append({"name": path.stem, "row_sample": min(len(rows), 100), "fields": fields})
    else:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover - environment guard
            raise RuntimeError("xlsx_requires_openpyxl") from exc
        workbook = load_workbook(path, read_only=True, data_only=False)
        for sheet in workbook.worksheets:
            values = sheet.iter_rows(values_only=True)
            headers = list(next(values, ()))
            sample_count = sum(1 for _ in values if _ is not None)
            profile["tables"].append({
                "name": sheet.title,
                "row_sample": min(sample_count, 100),
                "fields": [str(value) for value in headers if value is not None],
            })
    profile["table_count"] = len(profile["tables"])
    profile["status"] = "ready" if profile["tables"] else "empty"
    return profile


def _mapper_cli() -> Path:
    return Path(__file__).parents[2] / "skills" / "bifrost-data-mapper-readonly" / "scripts" / "mapper_cli.py"


FALLBACK_FIELD_ALIASES = {
    "LineID": "line_id", "line_id": "line_id", "产线": "line_id", "产线名称": "line_name",
    "班次日期": "shift_date", "生产日期": "shift_date", "日期": "shift_date",
    "shift_id": "shift_id", "shift": "shift_id", "班次": "shift_id", "班次号": "shift_id",
    "work_order_id": "work_order_id", "工单号": "work_order_id", "工单": "work_order_id",
    "product_id": "product_id", "product_code": "product_id", "产品编码": "product_id", "产品型号": "product_id",
    "equipment_id": "equipment_id", "machine_id": "equipment_id", "设备编号": "equipment_id", "设备ID": "equipment_id",
    "material_id": "material_id", "material_code": "material_id", "物料编码": "material_id", "物料号": "material_id",
    "defect_type": "defect_type", "defect_code": "defect_type", "不良类型": "defect_type", "缺陷类型": "defect_type",
    "stop_reason": "stop_reason", "reason_code": "stop_reason", "停机原因": "stop_reason", "停机类别": "stop_reason",
    "process_step": "process_step", "operation": "process_step", "工序": "process_step", "工艺": "process_step",
    "downtime_minutes": "downtime_minutes", "duration_minutes": "downtime_minutes", "stop_minutes": "downtime_minutes", "停机时长": "downtime_minutes", "停机分钟": "downtime_minutes",
    "defect_count": "defect_count", "defect_qty": "defect_count", "bad_qty": "defect_count", "不良数": "defect_count", "不良数量": "defect_count",
    "开动率": "availability", "availability": "availability",
    "性能率": "performance_rate_raw", "性能率(%)": "performance_rate_raw", "performance_rate": "performance_rate_raw",
    "质量率": "quality_rate", "良率(%)": "quality_rate", "良率": "quality_rate", "quality_rate": "quality_rate",
    "总产量": "total_output", "实际产量": "total_output", "实际数量": "total_output", "total_output": "total_output",
    "良品数": "good_output", "良品数量": "good_output", "good_output": "good_output",
    "OEE(%)": "oee_source", "OEE原值": "oee_source", "OEE": "oee_source", "oee": "oee_source",
}


def _fallback_mapping(path: Path, source_id: str, reason: str) -> dict[str, Any]:
    """Build a conservative local draft when the optional mapper is unavailable.

    Only exact, well-known header aliases are proposed.  Unknown columns are
    omitted, the overall status remains needs_confirmation, and no KPI is
    claimed to be approved by this fallback.
    """
    profile = _profile(path)
    draft = []
    for table in profile.get("tables", []):
        for field in table.get("fields", []):
            target = FALLBACK_FIELD_ALIASES.get(str(field))
            if not target:
                continue
            draft.append({
                "source_table": table["name"], "source_field": str(field),
                "target_entity": "production_shift", "target_field": target,
                "mapping_status": "proposed", "confidence": 0.70,
                "mapping_method": "deterministic_schema_alias_fallback",
                "fallback_requires_confirmation": True,
                "requires_human_confirmation": True,
            })
    return {
        "status": "needs_confirmation",
        "source_identity": {"detected_source_family": "LOCAL_DETERMINISTIC_FALLBACK"},
        "mapping_summary": {"confirmed": len(draft), "proposed": 0, "unmapped": 0, "fallback": 1},
        "mapping_draft": draft,
        "join_candidates": [],
        "data_quality_findings": [],
        "data_gaps": [reason],
        "fallback_used": True,
    }


def _run_mapper(path: Path, source_id: str) -> dict[str, Any]:
    """Call the existing mapper Skill through its read-only CLI."""
    cli = _mapper_cli()
    if not cli.exists():
        return {"status": "blocked", "data_gaps": ["mapper_skill_missing"]}
    with tempfile.NamedTemporaryFile(suffix=".json", delete=False) as output:
        output_path = Path(output.name)
    command = [
        sys.executable, str(cli), "--source-file", str(path), "--file-format", _format_for(path),
        "--request-id", f"REQ-AUTO-{source_id}", "--source-id", source_id, "--mapping-mode", "zero_shot",
        "--read-only", "true", "--output", str(output_path),
    ]
    try:
        # The mapper emits UTF-8 JSON/diagnostics.  Do not let the Windows
        # process default (often GBK) turn a valid source into a false
        # mapper_execution_failed result.
        completed = subprocess.run(
            command,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=180,
        )
        if output_path.exists() and output_path.stat().st_size:
            return json.loads(output_path.read_text(encoding="utf-8"))
        # Let ``run`` build the profile-aware deterministic draft.  Returning
        # an explicit blocked response here keeps the provider failure visible
        # and avoids the legacy fallback that marked aliases as confirmed.
        return {"status": "blocked", "data_gaps": ["mapper_no_response"], "stderr": completed.stderr[-1000:]}
    except (OSError, subprocess.SubprocessError, json.JSONDecodeError) as exc:
        return {"status": "blocked", "data_gaps": ["mapper_execution_failed"], "error": str(exc)}
    finally:
        output_path.unlink(missing_ok=True)


def _capabilities(mapping: dict[str, Any]) -> dict[str, Any]:
    mappings = mapping.get("mapping_draft", []) if isinstance(mapping, dict) else []
    confirmed = {
        item.get("target_field")
        for item in mappings
        if isinstance(item, dict) and item.get("target_field") and item.get("mapping_status") in {"confirmed", "proposed"}
        and float(item.get("confidence") or 0) >= 0.75
    }
    if "performance_rate_raw" in confirmed:
        confirmed.add("performance_rate")
    if "quality_factor" in confirmed:
        confirmed.add("quality_rate")
    result: dict[str, Any] = {}
    for name, required in REQUIRED_FIELDS.items():
        missing = sorted(required - confirmed)
        result[name] = {
            "status": "available" if not missing else "not_observable",
            "required_fields": sorted(required),
            "missing_fields": missing,
            "calculation_allowed": not missing,
        }
    return result


class AutoAdaptPipeline:
    """One source in, one safe adaptation contract out."""

    def run(self, source_file: str | Path, source_id: str = "SOURCE-LOCAL-001", confirmations: list[str] | None = None, drilldown_filters: dict[str, Any] | None = None) -> dict[str, Any]:
        path = Path(source_file).resolve()
        if not path.exists():
            raise FileNotFoundError(path)
        profile = _profile(path)
        source_id = _safe_id(source_id)
        mapping = _run_mapper(path, source_id)
        # An unavailable OMP mapper must not make the endpoint unusable, but a
        # local fallback may only provide a reviewable field draft.  It cannot
        # materialize canonical rows or produce KPI values without explicit
        # confirmation through the normal manifest gate.
        if mapping.get("status") == "blocked" and "mapper_no_response" in (mapping.get("data_gaps") or []):
            mapping = _deterministic_mapping_draft(profile)
        workstreams = Path(__file__).parents[1]
        if str(workstreams) not in sys.path:
            sys.path.insert(0, str(workstreams))
        omp_root = Path(__file__).parents[2]
        if str(omp_root) not in sys.path:
            sys.path.insert(0, str(omp_root))
        integration_root = omp_root / "integration"
        if str(integration_root) not in sys.path:
            sys.path.insert(0, str(integration_root))
        from compile.canonical_dataset import materialize_canonical_dataset
        from compile.mapping_confirmation import build_mapping_manifest
        from compile.payload_compiler import compile_payloads
        from drilldown import build_drilldown_manifest
        from governance_precheck import build_governance_report

        mapping_manifest = build_mapping_manifest(mapping, confirmations)
        # A partial mapping is a review draft, not a consumable dataset.  Do
        # not materialize rows or calculate KPIs until every required mapping
        # has been explicitly confirmed.  This prevents an API consumer from
        # accidentally treating a mixed/guessed schema as a valid result.
        if mapping_manifest.get("status") == "approved":
            canonical = materialize_canonical_dataset(path, mapping_manifest)
        else:
            canonical = {
                "contract_version": "BIFROST_CANONICAL_DATASET_v1",
                "source_sha256": profile["source_sha256"],
                "records": [],
                "record_count": 0,
                "evidence_index": [],
                "truncation": {"truncated": False, "reason": "mapping_not_approved"},
                "source_write_performed": False,
            }
        compiled_payloads = compile_payloads({"file_name": path.name, "format": _format_for(path), "source_sha256": profile["source_sha256"]}, mapping_manifest, canonical)
        drilldown_manifest = build_drilldown_manifest(
            canonical.get("records", []) if mapping_manifest.get("status") == "approved" else [],
            source_sha256=profile.get("source_sha256"),
        )
        # Keep active comparison lines separate from source-lineage identifiers.
        # The UI should offer only lines that the compiled payload can actually
        # query; lineage-only IDs remain available as audit metadata.
        coverage = compiled_payloads["overview"].get("view_coverage") or {}
        drilldown_manifest["active_line_ids"] = list(coverage.get("lines") or [])
        drilldown_manifest["source_line_ids"] = list(coverage.get("source_line_ids") or [])
        compiled_payloads["overview"]["drilldown_manifest"] = drilldown_manifest
        compiled_payloads["event"]["drilldown_manifest"] = drilldown_manifest
        # Use the same capability calculation that drives the compiled payload.
        # Keeping a second, raw-mapper-only calculation here caused the public
        # manifest to disagree with the payload (for example, performance_rate
        # was present in canonical rows but still reported as missing).
        capabilities = compiled_payloads["overview"].get("capability_manifest", {})
        # The manifest is the source of truth after an explicit confirmation.
        # Capability gaps (for example SPC/MTBF unavailable) are reported in
        # ``capability_manifest`` and must not make an already-approved
        # mapping appear to still require confirmation.
        effective_mapping_status = mapping_manifest["status"] if mapping_manifest.get("entries") else mapping.get("status", "blocked")
        needs_confirmation = effective_mapping_status != "approved"
        governance_report = build_governance_report(
            source_profile=profile,
            rows=canonical.get("records", []) if effective_mapping_status == "approved" else [],
            mapping_status=effective_mapping_status,
            capability_manifest=capabilities,
        )
        compiled_payloads["overview"]["data_quality_summary"] = governance_report
        compiled_payloads["event"]["data_quality_summary"] = governance_report
        result = {
            "contract_version": "BIFROST_AUTO_ADAPT_v1",
            "source_profile": profile,
            "mapping_summary": mapping.get("mapping_summary", {}),
            "mapping_status": effective_mapping_status,
            "quality_summary": {
                "finding_count": len(mapping.get("data_quality_findings", []) or []),
                "data_gaps": mapping.get("data_gaps", []),
            },
            "relation_summary": {
                "candidate_count": len(mapping.get("join_candidates", []) or []),
                "candidates": mapping.get("join_candidates", [])[:20],
            },
            "capability_manifest": capabilities,
            "governance_report": governance_report,
            "mapping_manifest": {
                "mapping_version": mapping_manifest["mapping_version"],
                "status": mapping_manifest["status"],
                "approved_count": mapping_manifest["approved_count"],
                "confirmation_count": mapping_manifest["confirmation_count"],
                "entry_count": len(mapping_manifest.get("entries", [])),
                "approved_mapping_ids": [
                    item.get("mapping_id") for item in mapping_manifest.get("entries", [])
                    if item.get("runtime_status") == "approved" and item.get("mapping_id")
                ],
                "pending_mapping_ids": [
                    item.get("mapping_id") for item in mapping_manifest.get("entries", [])
                    if item.get("runtime_status") != "approved" and item.get("mapping_id")
                ],
                "preview": [
                    {
                        "mapping_id": item.get("mapping_id"),
                        "source_table": item.get("source_table"),
                        "source_field": item.get("source_field"),
                        "target_field": item.get("target_field"),
                        "runtime_status": item.get("runtime_status"),
                        "requires_human_confirmation": item.get("requires_human_confirmation"),
                    }
                    for item in mapping_manifest.get("entries", [])[:50]
                ],
            },
            "canonical_dataset": {
                "status": "materialized_read_only",
                "record_count": canonical["record_count"],
                "source_sha256": canonical["source_sha256"],
                "source_write_performed": False,
            },
            "evidence_index": {"status": "physical", "count": len(canonical["evidence_index"]), "refs": canonical["evidence_index"][:20]},
            "needs_confirmation": bool(needs_confirmation),
            "data_gaps": mapping.get("data_gaps", []) or [],
            "source_write_performed": False,
            "drilldown_manifest": drilldown_manifest,
        }
        if mapping_manifest.get("status") == "approved" and drilldown_filters is not None:
            from drilldown import query_drilldown
            result["drilldown_result"] = query_drilldown(
                canonical.get("records", []),
                filters=drilldown_filters,
                source_sha256=profile.get("source_sha256"),
            )
        result["generated_payloads"] = compiled_payloads
        return result


__all__ = ["AutoAdaptPipeline"]
