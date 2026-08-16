"""Materialize a bounded canonical row set with physical evidence refs."""

from __future__ import annotations

import csv
import hashlib
import json
from pathlib import Path
from typing import Any


# Different source families may name the same business dimension differently.
# This is an internal semantic normalization, not a claim that the source
# field itself was renamed.
CANONICAL_TARGET_ALIASES = {
    # Keep a display name separate from the stable production-line key. A
    # name such as “设备瓶颈线” is not a valid identifier and must not leak
    # into line_id or be used for joins.
    "line_name": "line_name",
    "production_line": "line_id",
    "line": "line_id",
    "production_date": "shift_date",
    "actual_qty": "total_output",
    "actual_output": "total_output",
    "yield_rate": "quality_rate",
}


def _canonical_target(target_field: str) -> str:
    return CANONICAL_TARGET_ALIASES.get(target_field, target_field)


def _read_tables(path: Path) -> list[tuple[str, list[str], list[list[Any]]]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            rows = list(csv.reader(handle))
        return [(path.stem, rows[0] if rows else [], rows[1:])]
    if suffix == ".json":
        value = json.loads(path.read_text(encoding="utf-8"))
        rows = value if isinstance(value, list) else value.get("rows", []) if isinstance(value, dict) else []
        fields = sorted({key for row in rows if isinstance(row, dict) for key in row})
        return [(path.stem, fields, [[row.get(field) for field in fields] for row in rows if isinstance(row, dict)])]
    from openpyxl import load_workbook
    workbook = load_workbook(path, read_only=True, data_only=True)
    tables = []
    for sheet in workbook.worksheets:
        values = list(sheet.iter_rows(values_only=True))
        tables.append((sheet.title, [str(x) for x in (values[0] if values else [])], values[1:]))
    return tables


def _coerce_scalar(target_field: str, value: Any) -> Any:
    """Parse numeric source cells while leaving source files untouched."""
    if not isinstance(value, str):
        return value
    text = value.strip().replace(",", "")
    if not text:
        return value
    numeric_fields = {
        "availability", "performance_rate_raw", "quality_rate", "oee_source",
        "yield_recompute", "total_output", "good_output", "defect_count",
        "defect_ratio", "downtime_minutes", "actual_changeover_minutes",
        "standard_changeover_minutes", "planned_production_sec",
        "unplanned_downtime_sec", "required_quantity", "available_quantity",
        "quantity", "amount",
    }
    if target_field not in numeric_fields:
        return value
    try:
        number = float(text.rstrip("%"))
    except ValueError:
        return value
    return int(number) if number.is_integer() else number


def materialize_canonical_dataset(path: str | Path, mapping_manifest: dict[str, Any], *, max_rows: int = 20000) -> dict[str, Any]:
    source = Path(path).resolve()
    source_hash = hashlib.sha256(source.read_bytes()).hexdigest()
    approved = {
        (item.get("source_table"), item.get("source_field")): item
        for item in mapping_manifest.get("entries", [])
        if item.get("runtime_status") == "approved"
    }
    records: list[dict[str, Any]] = []
    evidence_index: list[dict[str, Any]] = []
    truncated = False
    for table_name, headers, rows in _read_tables(source):
        for row_number, row in enumerate(rows, start=2):
            if len(records) >= max_rows:
                truncated = True
                break
            record: dict[str, Any] = {"source_table": table_name, "source_row": row_number}
            row_evidence = f"SRC-{source_hash[:12]}:{table_name}:{row_number}"
            for index, header in enumerate(headers):
                mapping = approved.get((table_name, header))
                if not mapping or index >= len(row):
                    continue
                target_field = _canonical_target(mapping["target_field"])
                record[target_field] = _coerce_scalar(target_field, row[index])
                # Public spreadsheets often store ratios as percentages (e.g.
                # 95.7) while BIFROST contracts use 0-1 ratios. Normalize only
                # the semantic ratio fields and keep an audit note; raw source
                # cells remain untouched.
                value = record.get(target_field)
                if target_field in {"availability", "performance_rate_raw", "quality_rate", "oee_source", "yield_recompute"} and isinstance(value, (int, float)) and value > 1:
                    record[target_field] = value / 100
                    record.setdefault("normalization_notes", []).append(f"{target_field}:percent_to_ratio")
            # Excel formula caches can be empty when read by openpyxl.  These
            # are deterministic derived fields, so calculate them only when
            # all inputs are present and preserve the derivation marker.
            derived: list[str] = []
            total = record.get("total_output")
            quality = record.get("quality_rate")
            if record.get("good_output") is None and isinstance(total, (int, float)) and isinstance(quality, (int, float)):
                record["good_output"] = round(total * quality)
                derived.append("good_output=ROUND(total_output*quality_rate,0)")
            if record.get("yield_recompute") is None and isinstance(total, (int, float)) and total:
                good = record.get("good_output")
                if isinstance(good, (int, float)):
                    record["yield_recompute"] = good / total
                    derived.append("yield_recompute=good_output/total_output")
            if record.get("oee_recomputed") is None:
                parts = [record.get("availability"), record.get("performance_rate_raw"), record.get("quality_rate")]
                if all(isinstance(value, (int, float)) for value in parts):
                    record["oee_recomputed"] = parts[0] * parts[1] * parts[2]
                    derived.append("oee_recomputed=availability*performance_rate_raw*quality_rate")
            if len(record) > 2:
                record["evidence_ref"] = row_evidence
                if derived:
                    record["derived_fields"] = derived
                records.append(record)
                evidence_entry = {"evidence_ref": row_evidence, "source_table": table_name, "source_row": row_number, "source_sha256": source_hash}
                if record.get("line_id") not in (None, ""):
                    evidence_entry["line_id"] = record["line_id"]
                evidence_index.append(evidence_entry)
    return {
        "contract_version": "BIFROST_CANONICAL_DATASET_v1",
        "source_sha256": source_hash,
        "mapping_version": mapping_manifest.get("mapping_version"),
        "records": records,
        "record_count": len(records),
        "evidence_index": evidence_index,
        "truncation": {"truncated": truncated, "max_rows": max_rows},
        "source_write_performed": False,
    }


__all__ = ["materialize_canonical_dataset"]
